from django.shortcuts import render, redirect
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import FileResponse
from django.template.loader import get_template
from django.contrib.auth.decorators import login_required
from index.decorators import no_autenticado, allowed_users
from .forms import GeneroFormSexo, FormEstadoC, FormTallaC, FormTallaP, FormTallaZ, FormGradoI, FormTipoPersonal, FormCargo, FormDepartamento, FormSedes
from .models import Sexo, EstadoCivil,TallasCamisa, TallasPantalon, TallasZapatos, Grado, TipoPersonal, Cargo, Departamento, Sedes, Personal, Sangre, Nacionalidad
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime

def personal(request):
    personales = Personal.objects.all()
    cargos = cargo()
    sexos = sexo()
    estados = estado()
    sangres = sangre()
    talla_camisa = tallaCamisa()
    talla_pantalon = tallaPantalon()
    talla_zapato = tallaZapatos()
    grado_instruccion = gradoInstruccion()
    tipo_personal = tipoPersonal()
    departamentos = departamento()
    sedes = sede()
    nacionalidades = nacionalidad()
    return render(request, "ingresoPersonal.html", {"personales": personales, "cargos": cargos, "sexos": sexos, "estados": estados, "sangres": sangres, "talla_camisa": talla_camisa, "talla_pantalon": talla_pantalon, "talla_zapato": talla_zapato, "grado_instruccion": grado_instruccion, "tipo_personal": tipo_personal, "departamentos": departamentos, "sedes": sedes, "nacionalidades": nacionalidades })

def cargo():
    cargos = Cargo.objects.all()
    return cargos

def sexo():
    sexos = Sexo.objects.all()
    return sexos

def estado():
    estados = EstadoCivil.objects.all()
    return estados

def sangre():
    sangres = Sangre.objects.all()
    return sangres

def tallaCamisa():
    talla_camisa = TallasCamisa.objects.all()
    return talla_camisa
    
def tallaPantalon():
    talla_pantalon = TallasPantalon.objects.all()
    return talla_pantalon
   
def tallaZapatos():
    talla_zapato = TallasZapatos.objects.all()
    return talla_zapato

def gradoInstruccion():
    grado_instruccion = Grado.objects.all()
    return grado_instruccion

def tipoPersonal():
    tipo_personal = TipoPersonal.objects.all()
    return tipo_personal

def departamento():
    departamentos = Departamento.objects.all()
    return departamentos

def sede():
    sedes = Sedes.objects.all()
    return sedes

def nacionalidad():
    nacionalidades = Nacionalidad.objects.all()
    return nacionalidades
    
def estadoCivil(request):
    if request.method =='POST':
        form = FormEstadoC(request.POST)
        if form.is_valid():
            form.save()
    else: 
            form = FormEstadoC()
    estadosCiviles=EstadoCivil.objects.all()
    return render(request, "mantenimientos/estadoCivil.html", {"estadosCiviles":estadosCiviles, "form":form})

def mantSexo(request):
    if request.method =='POST':
        form = GeneroFormSexo(request.POST)
        if form.is_valid():
            form.save()
    else: 
            form = GeneroFormSexo()
    genero=Sexo.objects.all()
    return render(request, "mantenimientos/sexo.html", {"genero":genero, "form":form})

def mantTallaCamisa(request):
    if request.method =='POST':
        form = FormTallaC(request.POST)
        if form.is_valid():
            form.save()
    else: 
            form = FormTallaC()
    talla_camisa=TallasCamisa.objects.all()
    return render(request, "mantenimientos/tallasCamisa.html", {"talla_camisa":talla_camisa, "form":form})

def mantTallaPantalon(request):
    if request.method =='POST':
        form = FormTallaP(request.POST)
        if form.is_valid():
            form.save()
    else: 
            form = FormTallaP()
    talla_pantalon=TallasPantalon.objects.all()
    return render(request, "mantenimientos/tallasPantalon.html", {"talla_pantalon":talla_pantalon, "form":form})

def mantTallaZapato(request):
    if request.method =='POST':
        form = FormTallaZ(request.POST)
        if form.is_valid():
            form.save()
    else: 
            form = FormTallaZ()
    talla_zapato=TallasZapatos.objects.all()
    return render(request, "mantenimientos/tallasZapato.html", {"talla_zapato":talla_zapato, "form": form})

def mantGrado(request):
    if request.method =='POST':
        form = FormGradoI(request.POST)
        if form.is_valid():
            form.save()
    else: 
            form = FormGradoI()
    grado_instruccion=Grado.objects.all()
    return render(request, "mantenimientos/gradoInstruccion.html", {"grado_instruccion":grado_instruccion, "form": form})

def mantTipoPersonal(request):
    if request.method =='POST':
        form = FormTipoPersonal(request.POST)
        if form.is_valid():
            form.save()
    else: 
            form = FormTipoPersonal()
    tipo_personal=TipoPersonal.objects.all()
    return render(request, "mantenimientos/tipoPersonal.html", {"tipo_personal":tipo_personal, "form": form})


def mantCargo(request):
    if request.method =='POST':
        form = FormCargo(request.POST)
        if form.is_valid():
            form.save()
    else: 
            form = FormCargo()
    cargo=Cargo.objects.all()
    return render(request, "mantenimientos/cargo.html", {"cargo":cargo, "form": form})

def mantDepartamento(request):
    if request.method =='POST':
        form = FormDepartamento(request.POST)
        if form.is_valid():
            form.save()
    else: 
            form = FormDepartamento()
    departamento=Departamento.objects.all()
    return render(request, "mantenimientos/departamento.html", {"departamento":departamento, "form": form})

def mantSedes(request):
    if request.method =='POST':
        form = FormSedes(request.POST)
        if form.is_valid():
            form.save()
    else: 
            form = FormSedes()
    sedes=Sedes.objects.all()
    return render(request, "mantenimientos/sedes.html", {"sedes":sedes, "form": form})

def personalRetirado(request):
    personales = Personal.objects.filter(estatus='RETIRADO')
    cargos = cargo()
    sexos = sexo()
    estados = estado()
    sangres = sangre()
    talla_camisa = tallaCamisa()
    talla_pantalon = tallaPantalon()
    talla_zapato = tallaZapatos()
    grado_instruccion = gradoInstruccion()
    tipo_personal = tipoPersonal()
    departamentos = departamento()
    sedes = sede()
    nacionalidades = nacionalidad()
    return render(request, "personalRetirado.html", {"personales": personales, "cargos": cargos, "sexos": sexos, "estados": estados, "sangres": sangres, "talla_camisa": talla_camisa, "talla_pantalon": talla_pantalon, "talla_zapato": talla_zapato, "grado_instruccion": grado_instruccion, "tipo_personal": tipo_personal, "departamentos": departamentos, "sedes": sedes, "nacionalidades": nacionalidades })


def personalVacaciones(request):
    personales = Personal.objects.filter(estatus='VACACIONES')
    cargos = cargo()
    sexos = sexo()
    estados = estado()
    sangres = sangre()
    talla_camisa = tallaCamisa()
    talla_pantalon = tallaPantalon()
    talla_zapato = tallaZapatos()
    grado_instruccion = gradoInstruccion()
    tipo_personal = tipoPersonal()
    departamentos = departamento()
    sedes = sede()
    nacionalidades = nacionalidad()
    return render(request, "personalVacaciones.html", {"personales": personales, "cargos": cargos, "sexos": sexos, "estados": estados, "sangres": sangres, "talla_camisa": talla_camisa, "talla_pantalon": talla_pantalon, "talla_zapato": talla_zapato, "grado_instruccion": grado_instruccion, "tipo_personal": tipo_personal, "departamentos": departamentos, "sedes": sedes, "nacionalidades": nacionalidades })

def descargar_excel(request):
    # Filtra los datos del modelo para generar el archivo Excel
    personales = Personal.objects.all()

    # Crea un archivo Excel en memoria
    wb = Workbook()
    ws = wb.active
    
    # Agrega el título en la primera fila
    ws.merge_cells('A1:N1')  # Mezcla las celdas de A1 a D1
    ws['A1'] = 'Personal Completo del 911'  # Agrega el texto del título
    ws['A1'].alignment = Alignment(horizontal='center')  # Centra el texto del título
    ws['A1'].font = Font(bold=True, color="0000FF")  # Establece el texto en negrita y color azul
    ws['A2'] = ''  # Agrega el texto del título

    ws.append(["Estatus", "Nombres", "Apellidos", "Nacionalidad", "Cedula", "Genero", "Fecha Nac", "Edad", "Telefono", "Estado Civil", "Conyugue", "CI Conyugue", "Tipo Sangre", "Discapacitado", "Camisa", "Pantalon", "Zapatos", "Direccion", "Nro Cuenta", "Email", "Grado Instruccion", "Estudias", "Comision", "Pnb", "Tipo", "Cargo", "Fecha Ingreso 911", "Fecha Ingreso APN", "Contratos", "Departamentos", "Niños < 12", "Edad", "Hijos > 13", "Edad", "Niña < 12", "Edad", "Hijos Discapacidad", "Edad", "Motivo Retiro", "Sede", "Fasmij", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Fecha Creacion", "Fecha Actualizacion"])  # Reemplaza con los nombres de tus campos
    # Obtiene las columnas y establece el ancho deseado
    columnas = ws.column_dimensions
    columnas['A'].width = 15
    columnas['B'].width = 20
    columnas['C'].width = 20
    columnas['D'].width = 20
    columnas['E'].width = 10
    columnas['F'].width = 15
    columnas['G'].width = 12
    columnas['H'].width = 5
    columnas['I'].width = 12
    columnas['J'].width = 12
    columnas['K'].width = 15
    columnas['L'].width = 10
    columnas['M'].width = 12
    columnas['N'].width = 12
    columnas['O'].width = 10
    columnas['P'].width = 10
    columnas['Q'].width = 10
    columnas['R'].width = 30
    columnas['S'].width = 25
    columnas['T'].width = 25
    columnas['U'].width = 17
    columnas['V'].width = 10
    columnas['W'].width = 10
    columnas['X'].width = 10
    columnas['Y'].width = 10
    columnas['Z'].width = 20
    columnas['AA'].width = 18
    columnas['AB'].width = 18
    columnas['AC'].width = 10
    columnas['AD'].width = 25
    columnas['AE'].width = 18
    columnas['AF'].width = 8
    columnas['AG'].width = 18
    columnas['AH'].width = 8
    columnas['AI'].width = 18
    columnas['AJ'].width = 8
    columnas['AK'].width = 18
    columnas['AL'].width = 8
    columnas['AM'].width = 25
    columnas['AN'].width = 15
    columnas['AO'].width = 8
    columnas['AP'].width = 15
    columnas['AQ'].width = 20
    columnas['AR'].width = 15
    columnas['AS'].width = 30
    columnas['AT'].width = 15
    columnas['AU'].width = 20
    columnas['AV'].width = 15
    columnas['AW'].width = 30
    columnas['AX'].width = 15
    columnas['AY'].width = 20
    columnas['AZ'].width = 15
    columnas['BA'].width = 30
    columnas['BB'].width = 20
    columnas['BC'].width = 20
    for dato in personales:
       discapacitado =  dato.discapacitado
       if (discapacitado ==  True):
           discapacitado = 'SI'
       else:
           discapacitado = 'NO'
       estudias =  dato.estudias
       if (estudias ==  True):
           estudias = 'SI'
       else:
           estudias = 'NO' 
       comision =  dato.comision_servicio
       if (comision ==  True):
           comision = 'SI'
       else:
           comision = 'NO' 
       pnb =  dato.pnb
       if (pnb ==  True):
           pnb = 'SI'
       else:
           pnb = 'NO' 
       fasmij =  dato.fasmij
       if (fasmij ==  True):
           fasmij = 'SI'
       else:
           fasmij = 'NO' 
       ws.append([dato.estatus, dato.nombres, dato.apellidos, dato.nacionalidad.nacionalidad, dato.cedula, dato.sexo.sexo, dato.fecha_nac, dato.edad, dato.telefono, dato.estado_civil.estado_civil, dato.conyugue, dato.cedula_conyugue, dato.tipo_sangre.tipo_sangre, discapacitado, dato.talla_camisa.talla_camisa, dato.talla_pantalon.talla_pantalon, dato.talla_zapato.talla_zapato, dato.direccion, dato.nro_cuenta, dato.email, dato.grado_instruccion.grado_instruccion, estudias, comision, pnb, dato.tipo_personal.tipo_personal, dato.cargo.cargo, dato.fecha_ingreso_911, dato.fecha_ingreso_apn, dato.contratos, dato.departamento.departamento, dato.niño_menor_12, dato.edades1, dato.hijos_13_18, dato.edades2, dato.niña_menor_12, dato.edades3, dato.hijos_discapacidad, dato.edades4, dato.motivo, dato.sede.sede, fasmij, dato.parentezco1, dato.beneficiario1, dato.cedula1, dato.direccion1, dato.parentezco2, dato.beneficiario2, dato.cedula2, dato.direccion2, dato.parentezco3, dato.beneficiario3, dato.cedula3, dato.direccion3])  # Reemplaza con los campos de tu modelo

    # Convierte el archivo Excel en memoria a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Devuelve el archivo Excel como respuesta
    return FileResponse(output, as_attachment=True, filename="datosPersonal.xlsx")

def descargar_excel_retirados(request):
    # Filtra los datos del modelo para generar el archivo Excel
    #personales = Personal.objects.all()
    personales = Personal.objects.filter(estatus='RETIRADO')
    # Crea un archivo Excel en memoria
    wb = Workbook()
    ws = wb.active
    # Agrega el título en la primera fila
    ws.merge_cells('A1:N1')  # Mezcla las celdas de A1 a D1
    ws['A1'] = 'Personal Retirado del 911'  # Agrega el texto del título
    ws['A1'].alignment = Alignment(horizontal='center')  # Centra el texto del título
    ws['A1'].font = Font(bold=True, color="0000FF")  # Establece el texto en negrita y color azul
    ws['A2'] = ''  # Agrega el texto del título
    ws.append(["Estatus", "Nombres", "Apellidos", "Nacionalidad", "Cedula", "Genero", "Fecha Nac", "Edad", "Telefono", "Estado Civil", "Conyugue", "CI Conyugue", "Tipo Sangre", "Discapacitado", "Camisa", "Pantalon", "Zapatos", "Direccion", "Nro Cuenta", "Email", "Grado Instruccion", "Estudias", "Comision", "Pnb", "Tipo", "Cargo", "Fecha Ingreso 911", "Fecha Ingreso APN", "Contratos", "Departamentos", "Niños < 12", "Edad", "Hijos > 13", "Edad", "Niña < 12", "Edad", "Hijos Discapacidad", "Edad", "Motivo Retiro", "Sede", "Fasmij", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Fecha Creacion", "Fecha Actualizacion"])  # Reemplaza con los nombres de tus campos
    # Obtiene las columnas y establece el ancho deseado
    columnas = ws.column_dimensions
    columnas['A'].width = 15
    columnas['B'].width = 20
    columnas['C'].width = 20
    columnas['D'].width = 20
    columnas['E'].width = 10
    columnas['F'].width = 15
    columnas['G'].width = 12
    columnas['H'].width = 5
    columnas['I'].width = 12
    columnas['J'].width = 12
    columnas['K'].width = 15
    columnas['L'].width = 10
    columnas['M'].width = 12
    columnas['N'].width = 12
    columnas['O'].width = 10
    columnas['P'].width = 10
    columnas['Q'].width = 10
    columnas['R'].width = 30
    columnas['S'].width = 25
    columnas['T'].width = 25
    columnas['U'].width = 17
    columnas['V'].width = 10
    columnas['W'].width = 10
    columnas['X'].width = 10
    columnas['Y'].width = 10
    columnas['Z'].width = 20
    columnas['AA'].width = 18
    columnas['AB'].width = 18
    columnas['AC'].width = 10
    columnas['AD'].width = 25
    columnas['AE'].width = 18
    columnas['AF'].width = 8
    columnas['AG'].width = 18
    columnas['AH'].width = 8
    columnas['AI'].width = 18
    columnas['AJ'].width = 8
    columnas['AK'].width = 18
    columnas['AL'].width = 8
    columnas['AM'].width = 25
    columnas['AN'].width = 15
    columnas['AO'].width = 8
    columnas['AP'].width = 15
    columnas['AQ'].width = 20
    columnas['AR'].width = 15
    columnas['AS'].width = 30
    columnas['AT'].width = 15
    columnas['AU'].width = 20
    columnas['AV'].width = 15
    columnas['AW'].width = 30
    columnas['AX'].width = 15
    columnas['AY'].width = 20
    columnas['AZ'].width = 15
    columnas['BA'].width = 30
    columnas['BB'].width = 20
    columnas['BC'].width = 20
    for dato in personales:
       discapacitado =  dato.discapacitado
       if (discapacitado ==  True):
           discapacitado = 'SI'
       else:
           discapacitado = 'NO'
       estudias =  dato.estudias
       if (estudias ==  True):
           estudias = 'SI'
       else:
           estudias = 'NO' 
       comision =  dato.comision_servicio
       if (comision ==  True):
           comision = 'SI'
       else:
           comision = 'NO' 
           
       pnb =  dato.pnb
       if (pnb ==  True):
           pnb = 'SI'
       else:
           pnb = 'NO' 
           
       fasmij =  dato.fasmij
       if (fasmij ==  True):
           fasmij = 'SI'
       else:
           fasmij = 'NO' 
       ws.append([dato.estatus, dato.nombres, dato.apellidos, dato.nacionalidad.nacionalidad, dato.cedula, dato.sexo.sexo, dato.fecha_nac, dato.edad, dato.telefono, dato.estado_civil.estado_civil, dato.conyugue, dato.cedula_conyugue, dato.tipo_sangre.tipo_sangre, discapacitado, dato.talla_camisa.talla_camisa, dato.talla_pantalon.talla_pantalon, dato.talla_zapato.talla_zapato, dato.direccion, dato.nro_cuenta, dato.email, dato.grado_instruccion.grado_instruccion, estudias, comision, pnb, dato.tipo_personal.tipo_personal, dato.cargo.cargo, dato.fecha_ingreso_911, dato.fecha_ingreso_apn, dato.contratos, dato.departamento.departamento, dato.niño_menor_12, dato.edades1, dato.hijos_13_18, dato.edades2, dato.niña_menor_12, dato.edades3, dato.hijos_discapacidad, dato.edades4, dato.motivo, dato.sede.sede, fasmij, dato.parentezco1, dato.beneficiario1, dato.cedula1, dato.direccion1, dato.parentezco2, dato.beneficiario2, dato.cedula2, dato.direccion2, dato.parentezco3, dato.beneficiario3, dato.cedula3, dato.direccion3])  # Reemplaza con los campos de tu modelo
    # Convierte el archivo Excel en memoria a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    # Devuelve el archivo Excel como respuesta
    return FileResponse(output, as_attachment=True, filename="datosPersonalRetirados.xlsx")

def descargar_excel_vacaciones(request):
    # Filtra los datos del modelo para generar el archivo Excel
    #personales = Personal.objects.all()
    personales = Personal.objects.filter(estatus='VACACIONES')
    # Crea un archivo Excel en memoria
    wb = Workbook()
    ws = wb.active
    # Agrega el título en la primera fila
    ws.merge_cells('A1:N1')  # Mezcla las celdas de A1 a D1
    ws['A1'] = 'Personal de Vacaciones del 911'  # Agrega el texto del título
    ws['A1'].alignment = Alignment(horizontal='center')  # Centra el texto del título
    ws['A1'].font = Font(bold=True, color="0000FF")  # Establece el texto en negrita y color azul
    ws['A2'] = ''  # Agrega el texto del título
    ws.append(["Estatus", "Nombres", "Apellidos", "Nacionalidad", "Cedula", "Genero", "Fecha Nac", "Edad", "Telefono", "Estado Civil", "Conyugue", "CI Conyugue", "Tipo Sangre", "Discapacitado", "Camisa", "Pantalon", "Zapatos", "Direccion", "Nro Cuenta", "Email", "Grado Instruccion", "Estudias", "Comision", "Pnb", "Tipo", "Cargo", "Fecha Ingreso 911", "Fecha Ingreso APN", "Contratos", "Departamentos", "Niños < 12", "Edad", "Hijos > 13", "Edad", "Niña < 12", "Edad", "Hijos Discapacidad", "Edad", "Motivo Retiro", "Sede", "Fasmij", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Parentezco", "Beneficiario", "Cedula", "Direccion", "Fecha Creacion", "Fecha Actualizacion"])  # Reemplaza con los nombres de tus campos
    # Obtiene las columnas y establece el ancho deseado
    columnas = ws.column_dimensions
    columnas['A'].width = 15
    columnas['B'].width = 20
    columnas['C'].width = 20
    columnas['D'].width = 20
    columnas['E'].width = 10
    columnas['F'].width = 15
    columnas['G'].width = 12
    columnas['H'].width = 5
    columnas['I'].width = 12
    columnas['J'].width = 12
    columnas['K'].width = 15
    columnas['L'].width = 10
    columnas['M'].width = 12
    columnas['N'].width = 12
    columnas['O'].width = 10
    columnas['P'].width = 10
    columnas['Q'].width = 10
    columnas['R'].width = 30
    columnas['S'].width = 25
    columnas['T'].width = 25
    columnas['U'].width = 17
    columnas['V'].width = 10
    columnas['W'].width = 10
    columnas['X'].width = 10
    columnas['Y'].width = 10
    columnas['Z'].width = 20
    columnas['AA'].width = 18
    columnas['AB'].width = 18
    columnas['AC'].width = 10
    columnas['AD'].width = 25
    columnas['AE'].width = 18
    columnas['AF'].width = 8
    columnas['AG'].width = 18
    columnas['AH'].width = 8
    columnas['AI'].width = 18
    columnas['AJ'].width = 8
    columnas['AK'].width = 18
    columnas['AL'].width = 8
    columnas['AM'].width = 25
    columnas['AN'].width = 15
    columnas['AO'].width = 8
    columnas['AP'].width = 15
    columnas['AQ'].width = 20
    columnas['AR'].width = 15
    columnas['AS'].width = 30
    columnas['AT'].width = 15
    columnas['AU'].width = 20
    columnas['AV'].width = 15
    columnas['AW'].width = 30
    columnas['AX'].width = 15
    columnas['AY'].width = 20
    columnas['AZ'].width = 15
    columnas['BA'].width = 30
    columnas['BB'].width = 20
    columnas['BC'].width = 20
    for dato in personales:
       discapacitado =  dato.discapacitado
       if (discapacitado ==  True):
           discapacitado = 'SI'
       else:
           discapacitado = 'NO'
       estudias =  dato.estudias
       if (estudias ==  True):
           estudias = 'SI'
       else:
           estudias = 'NO' 
       comision =  dato.comision_servicio
       if (comision ==  True):
           comision = 'SI'
       else:
           comision = 'NO' 
       pnb =  dato.pnb
       if (pnb ==  True):
           pnb = 'SI'
       else:
           pnb = 'NO' 
       fasmij =  dato.fasmij
       if (fasmij ==  True):
           fasmij = 'SI'
       else:
           fasmij = 'NO' 
       ws.append([dato.estatus, dato.nombres, dato.apellidos, dato.nacionalidad.nacionalidad, dato.cedula, dato.sexo.sexo, dato.fecha_nac, dato.edad, dato.telefono, dato.estado_civil.estado_civil, dato.conyugue, dato.cedula_conyugue, dato.tipo_sangre.tipo_sangre, discapacitado, dato.talla_camisa.talla_camisa, dato.talla_pantalon.talla_pantalon, dato.talla_zapato.talla_zapato, dato.direccion, dato.nro_cuenta, dato.email, dato.grado_instruccion.grado_instruccion, estudias, comision, pnb, dato.tipo_personal.tipo_personal, dato.cargo.cargo, dato.fecha_ingreso_911, dato.fecha_ingreso_apn, dato.contratos, dato.departamento.departamento, dato.niño_menor_12, dato.edades1, dato.hijos_13_18, dato.edades2, dato.niña_menor_12, dato.edades3, dato.hijos_discapacidad, dato.edades4, dato.motivo, dato.sede.sede, fasmij, dato.parentezco1, dato.beneficiario1, dato.cedula1, dato.direccion1, dato.parentezco2, dato.beneficiario2, dato.cedula2, dato.direccion2, dato.parentezco3, dato.beneficiario3, dato.cedula3, dato.direccion3])  # Reemplaza con los campos de tu modelo
    # Convierte el archivo Excel en memoria a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    # Devuelve el archivo Excel como respuesta
    return FileResponse(output, as_attachment=True, filename="datosPersonalVacaciones.xlsx")
