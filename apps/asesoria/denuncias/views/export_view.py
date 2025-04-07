from io import BytesIO
from asesoria.denuncias.models import Denuncia, ESTATUS_CHOICES
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


class DenunciaExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "asesoria.denuncias.exel_denuncia"

    def get(self, request, *args, **kwargs):
        denuncias = Denuncia.objects.all().values(
            "estatus",
            "ente",
            "denunciante__nombres",
            "denunciante__apellidos",
            "denunciante__cedula",
            "denunciante__telefono",
            "denunciante__email",
            "denunciante__direccion",
            "denunciado__nombres",
            "denunciado__apellidos",
            "denunciado__cedula",
            "motivo",
            "zona",
            "fecha_denuncia",
            "fecha_incidente",
        )

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:P1")
        ws["A1"] = "Denuncias del 911"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF")
        ws["A2"] = ""

        # Crear mapeo de estatus
        estatus_mapping = dict(ESTATUS_CHOICES)

        # Encabezados
        ws.append(
            [
                "Estatus",
                "Ente",
                "Nombres del denunciante",
                "Apellidos del denunciante",
                "Cedula del denunciante",
                "Telefono",
                "Email",
                "Direccion",
                "Nombres del denunciado",
                "Apellidos del denunciado",
                "Cedula del denunciado",
                "Motivo",
                "Zona",
                "Fecha de la denuncia",
                "Fecha del incidente",
            ]
        )

        # Configuración de ancho de columnas
        columnas = ws.column_dimensions
        columnas["A"].width = 10
        columnas["B"].width = 15
        columnas["C"].width = 20
        columnas["D"].width = 20
        columnas["E"].width = 25
        columnas["F"].width = 25
        columnas["G"].width = 25
        columnas["H"].width = 25
        columnas["I"].width = 25
        columnas["J"].width = 25
        columnas["K"].width = 25
        columnas["L"].width = 25
        columnas["M"].width = 25
        columnas["N"].width = 25
        columnas["O"].width = 20
        columnas["P"].width = 20

        # Datos
        for dato in denuncias:
            # Formatear el estatus usando el mapeo
            estatus = estatus_mapping.get(dato["estatus"], dato["estatus"])

            ws.append(
                [
                    estatus,  # Estatus formateado
                    dato["ente"],
                    dato["denunciante__nombres"],
                    dato["denunciante__apellidos"],
                    dato["denunciante__cedula"],
                    dato["denunciante__telefono"],
                    dato["denunciante__email"],
                    dato["denunciante__direccion"],
                    dato["denunciado__nombres"],
                    dato["denunciado__apellidos"],
                    dato["denunciado__cedula"],
                    dato["motivo"],
                    dato["zona"],
                    dato["fecha_denuncia"],
                    dato["fecha_incidente"],
                ]
            )

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="Denuncias.xlsx")
