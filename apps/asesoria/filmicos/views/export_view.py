from io import BytesIO
from asesoria.filmicos.models import RegistroFilmico, ESTATUS_CHOICES
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


class RegistroFilmicoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "asesoria.filmicos.exel_registroFilmico"

    def get(self, request, *args, **kwargs):
        # Filtra los datos del modelo para generar el archivo Excel
        registros = RegistroFilmico.objects.all()

        # Crear mapeo de estatus
        estatus_mapping = dict(ESTATUS_CHOICES)

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:G1")
        ws["A1"] = "Registros Fílmicos del 911"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF")
        ws["A2"] = ""

        # Encabezados
        ws.append(
            [
                "Estatus",
                "Direccion",
                "Camara",
                "Motivo de solicitud",
                "Ente que solicita",
                "Fecha de la solicitud",
                "Fecha de culminacion",
            ]
        )

        # Configuración de ancho de columnas
        columnas = ws.column_dimensions
        columnas["A"].width = 15
        columnas["B"].width = 20
        columnas["C"].width = 20
        columnas["D"].width = 20
        columnas["E"].width = 30
        columnas["F"].width = 30
        columnas["G"].width = 30

        # Datos
        for registro in registros:
            # Formatear el estatus usando el mapeo
            estatus = estatus_mapping.get(registro.estatus, registro.estatus)

            ws.append(
                [
                    estatus,  # Estatus formateado
                    registro.direccion,
                    registro.camara,
                    registro.motivo_solicitud,
                    registro.ente_solicita,
                    registro.fecha_solicitud,
                    registro.fecha_culminacion,
                ]
            )

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="RegistroFilmico.xlsx")
