from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from emergencia.incidencias.models import TipoIncidencia 

class IncidenciasExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "emergencia.listar_tipo_incidencia_emergencia"  # Permiso del modelo
    
    def get(self, request, *args, **kwargs):
        # Consulta optimizada para TipoIncidencia
        incidencias = TipoIncidencia.objects.all().values(
            'id',
            'nombre_incidencia',
            'created_at',
            'updated_at',
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Tipos de Incidencia"

        # Configuración del título
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = "REGISTRO DE TIPOS DE INCIDENCIA"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados (ajustados a los campos del modelo)
        headers = [
            ("ID", 10),
            ("Nombre Incidencia", 30),
            ("Fecha Creación", 20),
            ("Creado En", 20),
        ]

        # Configurar columnas
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Llenar datos
        for row_num, incidencia in enumerate(incidencias, 3):
            ws.append([
                incidencia['id'],
                incidencia['nombre_incidencia'],
                incidencia['created_at'].strftime("%d/%m/%Y %H:%M") if incidencia['created_at'] else "N/A",
                incidencia['updated_at'].strftime("%d/%m/%Y %H:%M") if incidencia['updated_at'] else "N/A",
            ])

        # Aplicar bordes
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output, 
            as_attachment=True, 
            filename="reporte_tipos_incidencia.xlsx"  # Nombre del archivo ajustado
        )