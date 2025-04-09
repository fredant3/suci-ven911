from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from emergencia.operaciones.models import Emergencia, ESTADOS_CHOICES


class EmergenciaExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "emergencia.listar_emergencia"

    def get(self, request, *args, **kwargs):
        # Consulta optimizada corregida
        emergencias = Emergencia.objects.select_related(
            "organismo", "incidencia"
        ).values(
            "denunciante",
            "telefono_denunciante",
            "estado",
            "municipio",
            "parroquia",
            "organismo__nombre",
            "incidencia__nombre_incidencia",  # Nombre CORRECTO del campo
            "direccion_incidencia",
            "observaciones",
            "telefono_cuadrante_paz",
            "created_at",
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Emergencias"

        # Configuración del título (ajustado a 11 columnas)
        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = "REGISTRO DE EMERGENCIAS"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados actualizados (11 columnas)
        headers = [
            ("Denunciante", 25),
            ("Teléfono Denunciante", 20),
            ("Estado", 15),
            ("Municipio", 20),
            ("Parroquia", 20),
            ("Organismo", 25),
            ("Tipo Incidencia", 25),
            ("Dirección", 30),
            ("Observaciones", 40),
            ("Teléfono Cuadrante", 20),
            ("Fecha Registro", 20),
        ]

        # Configurar columnas
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Mapeo de estados
        estados_map = dict(ESTADOS_CHOICES)

        # Llenar datos con campo corregido
        for row_num, emergencia in enumerate(emergencias, 3):
            ws.append(
                [
                    emergencia["denunciante"],
                    emergencia["telefono_denunciante"] or "N/A",
                    estados_map.get(emergencia["estado"], "Desconocido"),
                    emergencia["municipio"],
                    emergencia["parroquia"],
                    emergencia["organismo__nombre"],
                    emergencia["incidencia__nombre_incidencia"],  # Campo CORRECTO aquí
                    emergencia["direccion_incidencia"] or "N/A",
                    emergencia["observaciones"] or "N/A",
                    emergencia["telefono_cuadrante_paz"] or "N/A",
                    (
                        emergencia["created_at"].strftime("%d/%m/%Y")
                        if emergencia["created_at"]
                        else "N/A"
                    ),
                ]
            )

        # Aplicar bordes
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output, as_attachment=True, filename="reporte_emergencias.xlsx"
        )
