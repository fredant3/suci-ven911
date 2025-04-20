from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from emergencia.organismo.models import Organismo

class OrganismoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "emergencia.listar_organismo_emergencia"  # Permiso corregido
    
    def get(self, request, *args, **kwargs):
        # Consulta optimizada para Organismo
        organismos = Organismo.objects.all().values(
            'id',
            'nombre',
            'created_at',
            'updated_at',
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Organismos"

        # Configuración del título
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "REGISTRO DE ORGANISMOS COMPETENTES"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados (ajustados a los campos del modelo Organismo)
        headers = [
            ("ID", 10),
            ("Nombre Organismo", 30),
            ("Fecha Creación", 20),
            ("Última Actualización", 20),
        ]

        # Configurar columnas
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Llenar datos
        for row_num, organismo in enumerate(organismos, 3):
            ws.append([
                organismo['id'],
                organismo['nombre'],
                organismo['created_at'].strftime("%d/%m/%Y %H:%M") if organismo['created_at'] else "N/A",
                organismo['updated_at'].strftime("%d/%m/%Y %H:%M") if organismo['updated_at'] else "N/A",
            ])

        # Aplicar bordes
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output, 
            as_attachment=True, 
            filename="reporte_organismos.xlsx"  # Nombre del archivo ajustado
        )