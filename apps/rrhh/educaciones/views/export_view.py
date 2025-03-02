from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from rrhh.educaciones.models import (
    Educacion,
)  # Asegúrate de importar el modelo Educacion


class EducacionExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = (
        "rrhh.educaciones.exel_educacion"  # Define el permiso necesario
    )

    def get(self, request, *args, **kwargs):
        educaciones = Educacion.objects.all().values(
            "colegio",
            "codigo_titulo",
            "titulo",
            "area_conocimiento",
            "fecha_inicio",
            "fecha_culminacion",
            "enlace_certificado",
            "empleado__nombres",  # Asume que el modelo Empleado tiene un campo 'nombres'
            "empleado__apellidos",  # Asume que el modelo Empleado tiene un campo 'apellidos'
        )

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:H1")  # Mezcla las celdas de A1 a H1
        ws["A1"] = "Listado de Educación"  # Agrega el texto del título
        ws["A1"].alignment = Alignment(
            horizontal="center"
        )  # Centra el texto del título
        ws["A1"].font = Font(
            bold=True, color="0000FF"
        )  # Establece el texto en negrita y color azul
        ws["A2"] = ""  # Agrega una fila vacía

        # Agrega los encabezados de las columnas
        ws.append(
            [
                "Colegio",
                "Código del Título",
                "Título",
                "Área de Conocimiento",
                "Fecha de Inicio",
                "Fecha de Culminación",
                "Enlace del Certificado",
                "Empleado",
            ]
        )

        # Ajusta el ancho de las columnas
        columnas = ws.column_dimensions
        columnas["A"].width = 30
        columnas["B"].width = 20
        columnas["C"].width = 30
        columnas["D"].width = 25
        columnas["E"].width = 20
        columnas["F"].width = 20
        columnas["G"].width = 30
        columnas["H"].width = 30

        # Agrega los datos de las educaciones
        for educacion in educaciones:
            ws.append(
                [
                    educacion["colegio"],
                    educacion["codigo_titulo"],
                    educacion["titulo"],
                    educacion["area_conocimiento"],
                    (
                        educacion["fecha_inicio"].strftime("%Y-%m-%d")
                        if educacion["fecha_inicio"]
                        else ""
                    ),
                    (
                        educacion["fecha_culminacion"].strftime("%Y-%m-%d")
                        if educacion["fecha_culminacion"]
                        else ""
                    ),
                    educacion["enlace_certificado"],
                    f"{educacion['empleado__nombres']} {educacion['empleado__apellidos']}",
                ]
            )

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="Educaciones.xlsx")
