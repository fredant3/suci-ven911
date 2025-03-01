from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from rrhh.dotaciones.models import Dotacion  # Asegúrate de importar el modelo Dotacion


class DotacionExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "rrhh.dotaciones.exel_dotacion"  # Define el permiso necesario

    def get(self, request, *args, **kwargs):
        dotaciones = Dotacion.objects.all().values(
            "camisa",
            "pantalon",
            "zapato",
            "empleado__nombres",  # Asume que el modelo Empleado tiene un campo 'nombres'
            "empleado__apellidos",  # Asume que el modelo Empleado tiene un campo 'apellidos'
        )

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:D1")  # Mezcla las celdas de A1 a D1
        ws["A1"] = "Listado de Dotaciones"  # Agrega el texto del título
        ws["A1"].alignment = Alignment(
            horizontal="center"
        )  # Centra el texto del título
        ws["A1"].font = Font(
            bold=True, color="0000FF"
        )  # Establece el texto en negrita y color azul
        ws["A2"] = ""  # Agrega una fila vacía

        # Agrega los encabezados de las columnas
        ws.append(
            [
                "Talla de Camisa",
                "Talla de Pantalón",
                "Talla de Zapato",
                "Empleado",
            ]
        )

        # Ajusta el ancho de las columnas
        columnas = ws.column_dimensions
        columnas["A"].width = 20
        columnas["B"].width = 20
        columnas["C"].width = 20
        columnas["D"].width = 30

        # Agrega los datos de las dotaciones
        for dotacion in dotaciones:
            ws.append(
                [
                    dotacion["camisa"],
                    dotacion["pantalon"],
                    dotacion["zapato"],
                    f"{dotacion['empleado__nombres']} {dotacion['empleado__apellidos']}",
                ]
            )

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="Dotaciones.xlsx")
