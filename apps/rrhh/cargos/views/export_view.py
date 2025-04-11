from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from rrhh.cargos.models import Cargo  # Asegúrate de importar el modelo Cargo


class CargoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "rrhh.cargo.exel_cargo"  # Define el permiso necesario

    def get(self, request, *args, **kwargs):
        cargos = Cargo.objects.all().values("cargo", "estatus", "created_at")

        # Diccionario para mapear las claves de estatus a sus valores descriptivos
        ESTATUS_MAP = {
            "act": "Activo",
            "ina": "Inactivo",
            "inv": "Invalido",
            "cer": "Cerrado",
        }

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:B1")  # Mezcla las celdas de A1 a B1
        ws["A1"] = "Listado de Cargos"  # Agrega el texto del título
        ws["A1"].alignment = Alignment(
            horizontal="center"
        )  # Centra el texto del título
        ws["A1"].font = Font(
            bold=True, color="0000FF"
        )  # Establece el texto en negrita y color azul
        ws["A2"] = ""  # Agrega una fila vacía

        # Agrega los encabezados de las columnas
        ws.append(["Cargo", "Estatus", "Fecha de creación"])

        # Ajusta el ancho de las columnas
        columnas = ws.column_dimensions
        columnas["A"].width = 30
        columnas["B"].width = 15

        # Agrega los datos de los cargos, convirtiendo la clave de estatus a su valor descriptivo
        for cargo in cargos:
            estatus_descriptivo = ESTATUS_MAP.get(cargo["estatus"], cargo["estatus"])
            ws.append(
                [
                    cargo["cargo"],
                    estatus_descriptivo,
                    (
                        cargo["created_at"].strftime("%d/%m/%Y")
                        if cargo["created_at"]
                        else "N/A"
                    ),
                ]
            )

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="Cargos.xlsx")
