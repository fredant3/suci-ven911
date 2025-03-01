from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from rrhh.sueldos.models import Sueldo  # Asegúrate de importar el modelo Sueldo


class SueldoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "rrhh.sueldos.exel_sueldo"  # Define el permiso necesario

    def get(self, request, *args, **kwargs):
        sueldos = Sueldo.objects.all().values(
            "tipo_sueldo__tipo",  # Asume que el modelo TipoSueldo tiene un campo 'tipo'
            "estatus",
            "fecha_pago",
            "monto",
            "empleado__nombres",  # Asume que el modelo Empleado tiene un campo 'nombres'
            "empleado__apellidos",  # Asume que el modelo Empleado tiene un campo 'apellidos'
        )

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:E1")  # Mezcla las celdas de A1 a E1
        ws["A1"] = "Listado de Sueldos"  # Agrega el texto del título
        ws["A1"].alignment = Alignment(
            horizontal="center"
        )  # Centra el texto del título
        ws["A1"].font = Font(
            bold=True, color="0000FF"
        )  # Establece el texto en negrita y color azul
        ws["A2"] = ""  # Agrega una fila vacía

        # Agrega los encabezados de las columnas
        ws.append(
            [
                "Tipo de Sueldo",
                "Estatus",
                "Fecha de Pago",
                "Monto",
                "Empleado",
            ]
        )

        # Ajusta el ancho de las columnas
        columnas = ws.column_dimensions
        columnas["A"].width = 20
        columnas["B"].width = 15
        columnas["C"].width = 15
        columnas["D"].width = 15
        columnas["E"].width = 30

        # Agrega los datos de los sueldos
        for sueldo in sueldos:
            ws.append(
                [
                    sueldo["tipo_sueldo__tipo"],
                    sueldo["estatus"],
                    (
                        sueldo["fecha_pago"].strftime("%Y-%m-%d")
                        if sueldo["fecha_pago"]
                        else ""
                    ),
                    float(
                        sueldo["monto"]
                    ),  # Convierte Decimal a float para evitar errores
                    f"{sueldo['empleado__nombres']} {sueldo['empleado__apellidos']}",
                ]
            )

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="Sueldos.xlsx")
