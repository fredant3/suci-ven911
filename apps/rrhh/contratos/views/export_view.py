from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from rrhh.contratos.models import (
    Contrato,
    TIPO_CONTRATOS_CHOICES,
    ESTATUS_CONTRATO_CHOICES,
)


class ContratoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    # Permiso requerido para acceder a esta vista (debe coincidir con los permisos del modelo)
    permission_required = "rrhh.contratos.exel_contrato"

    def get(self, request, *args, **kwargs):
        # Obtener todos los contratos de la base de datos
        contratos = Contrato.objects.all()

        # Crear diccionarios para mapear los valores de selección a sus representaciones legibles
        tipo_mapping = dict(TIPO_CONTRATOS_CHOICES)  # Mapeo para tipos de contrato
        estatus_mapping = dict(
            ESTATUS_CONTRATO_CHOICES
        )  # Mapeo para estatus de contrato
        boolean_mapping = {True: "Sí", False: "No"}  # Mapeo para valores booleanos

        # Crear un nuevo libro de Excel
        wb = Workbook()
        ws = wb.active  # Obtener la hoja activa

        # Configurar el título del reporte
        ws.merge_cells("A1:K1")  # Combinar celdas para el título
        ws["A1"] = "Registro de Contratos"  # Texto del título
        ws["A1"].alignment = Alignment(horizontal="center")  # Centrar el título
        ws["A1"].font = Font(
            bold=True, color="0000FF"
        )  # Formato del título (negrita y azul)
        ws.append([])  # Agregar una fila vacía después del título

        # Definir los encabezados de las columnas
        headers = [
            "Empleado",
            "Tipo de Contrato",
            "Comisión de Servicio",
            "Funcionario PNB",
            "Departamento",
            "Cargo Asignado",
            "Sede",
            "Fecha Ingreso Ven-911",
            "Fecha Ingreso APN",
            "Fecha Ingreso",
            "Fecha Culminación",
            "Estatus",
        ]
        ws.append(headers)  # Agregar los encabezados a la hoja

        # Configurar los anchos de las columnas para mejor visualización
        column_widths = {
            "A": 30,  # Empleado
            "B": 20,  # Tipo de Contrato
            "C": 20,  # Comisión de Servicio
            "D": 18,  # Funcionario PNB
            "E": 25,  # Departamento
            "F": 25,  # Cargo Asignado
            "G": 25,  # Sede
            "H": 20,  # Fecha Ingreso Ven-911
            "I": 20,  # Fecha Ingreso APN
            "J": 15,  # Fecha Ingreso
            "K": 20,  # Fecha Culminación
            "L": 15,  # Estatus
        }
        # Aplicar los anchos de columna configurados
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Llenar las filas con los datos de los contratos
        for contrato in contratos:
            ws.append(
                [
                    str(
                        contrato.empleado
                    ),  # Nombre del empleado (usa __str__ del modelo Empleado)
                    tipo_mapping.get(
                        contrato.tipo, contrato.tipo
                    ),  # Tipo de contrato (mapeado)
                    boolean_mapping.get(
                        contrato.comision_servicio, "No"
                    ),  # Comisión servicio (Sí/No)
                    boolean_mapping.get(contrato.pnb, "No"),  # Funcionario PNB (Sí/No)
                    (
                        str(contrato.departamento) if contrato.departamento else ""
                    ),  # Departamento
                    str(contrato.cargo) if contrato.cargo else "",  # Cargo asignado
                    str(contrato.sede) if contrato.sede else "",  # Sede
                    contrato.fecha_ingreso_911,  # Fecha ingreso Ven-911
                    contrato.fecha_ingreso_apn,  # Fecha ingreso APN
                    contrato.fecha_ingreso,  # Fecha de ingreso
                    (
                        contrato.fecha_culminacion if contrato.fecha_culminacion else ""
                    ),  # Fecha culminación (maneja nulos)
                    estatus_mapping.get(
                        contrato.estatus, contrato.estatus
                    ),  # Estatus del contrato (mapeado)
                ]
            )

        # Guardar el libro de Excel en memoria
        output = BytesIO()
        wb.save(output)
        output.seek(0)  # Rebobinar el buffer al inicio

        # Devolver el archivo Excel como respuesta para descargar
        return FileResponse(
            output,
            as_attachment=True,  # Forzar descarga
            filename="RegistroContratos.xlsx",  # Nombre del archivo
        )
