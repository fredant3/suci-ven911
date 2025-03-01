from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from rrhh.familiares.models import Familiar  # Asegúrate de importar el modelo Familiar


class FamiliarExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = (
        "rrhh.familiares.exportar_familiares"  # Define el permiso necesario
    )

    def get(self, request, *args, **kwargs):
        familiares = Familiar.objects.all().values(
            "parentezco",
            "tipo_hijo",
            "discapacidad",
            "nombres",
            "apellidos",
            "cedula",
            "fecha_nacimiento",
            "sexo",
            "estado_civil",
            "empleado__nombres",  # Asume que el modelo Empleado tiene un campo 'nombres'
            "empleado__apellidos",  # Asume que el modelo Empleado tiene un campo 'apellidos'
            "observacion",
        )

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:K1")  # Mezcla las celdas de A1 a K1
        ws["A1"] = "Listado de Familiares"  # Agrega el texto del título
        ws["A1"].alignment = Alignment(
            horizontal="center"
        )  # Centra el texto del título
        ws["A1"].font = Font(
            bold=True, color="0000FF"
        )  # Establece el texto en negrita y color azul
        ws["A2"] = ""  # Agrega una fila vacía

        # Agrega los encabezados de las columnas
        ws.append(
            [
                "Parentezco",
                "Tipo de Hijo",
                "Discapacidad",
                "Nombres",
                "Apellidos",
                "Cédula",
                "Fecha de Nacimiento",
                "Sexo",
                "Estado Civil",
                "Empleado",
                "Observación",
            ]
        )

        # Ajusta el ancho de las columnas
        columnas = ws.column_dimensions
        columnas["A"].width = 15
        columnas["B"].width = 15
        columnas["C"].width = 15
        columnas["D"].width = 20
        columnas["E"].width = 20
        columnas["F"].width = 15
        columnas["G"].width = 20
        columnas["H"].width = 10
        columnas["I"].width = 15
        columnas["J"].width = 30
        columnas["K"].width = 30

        # Agrega los datos de los familiares
        for familiar in familiares:
            ws.append(
                [
                    familiar["parentezco"],
                    familiar["tipo_hijo"] if familiar["tipo_hijo"] else "",
                    "Sí" if familiar["discapacidad"] == "yes" else "No",
                    familiar["nombres"],
                    familiar["apellidos"],
                    familiar["cedula"] if familiar["cedula"] else "",
                    (
                        familiar["fecha_nacimiento"].strftime("%Y-%m-%d")
                        if familiar["fecha_nacimiento"]
                        else ""
                    ),
                    familiar["sexo"],
                    familiar["estado_civil"],
                    f"{familiar['empleado__nombres']} {familiar['empleado__apellidos']}",
                    familiar["observacion"] if familiar["observacion"] else "",
                ]
            )

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="Familiares.xlsx")
