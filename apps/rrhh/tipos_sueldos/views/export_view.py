from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from rrhh.tipos_sueldos.models import (
    TipoSueldo,
)  # Asegúrate de importar el modelo TipoSueldo


class TipoSueldoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = (
        "rrhh.tipos_sueldos.exel_tipo_sueldo"  # Define el permiso necesario
    )

    def get(self, request, *args, **kwargs):
        tipos_sueldos = TipoSueldo.objects.all().values(
            "tipo",
            "monto",
            "descripcion",
            "estatus",
        )

        # Diccionarios para mapear las claves a sus valores descriptivos
        TIPO_MAP = {
            "ticket": "Cesta Ticket",
            "guerra": "Bono de Guerra",
            "discapacidad": "Prima por discapacidad",
            "menor_12": "Prima por dependencias menores de 12",
            "hijos_13_18": "Prima por dependencias menores de 13 a 18",
            "hijos_discapacidad": "Prima por dependencias menores con discapacidad",
            "profesionalismo": "Prima por Profesionalismo",
            "minimo": "Sueldo Mínimo",
        }

        ESTATUS_MAP = {
            "act": "Activo",
            "sup": "Suspendido",
        }

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:D1")  # Mezcla las celdas de A1 a D1
        ws["A1"] = "Listado de Tipos de Sueldos"  # Agrega el texto del título
        ws["A1"].alignment = Alignment(
            horizontal="center"
        )  # Centra el texto del título
        ws["A1"].font = Font(
            bold=True, color="0000FF"
        )  # Establece el texto en negrita y color azul
        ws["A2"] = ""  # Agrega una fila vacía

        # Agrega los encabezados de las columnas
        ws.append(
            [
                "Tipo de Sueldo",
                "Monto",
                "Descripción",
                "Estatus",
            ]
        )

        # Ajusta el ancho de las columnas
        columnas = ws.column_dimensions
        columnas["A"].width = 25
        columnas["B"].width = 15
        columnas["C"].width = 30
        columnas["D"].width = 15

        # Agrega los datos de los tipos de sueldos, convirtiendo las claves a valores descriptivos
        for tipo_sueldo in tipos_sueldos:
            tipo_descriptivo = TIPO_MAP.get(tipo_sueldo["tipo"], tipo_sueldo["tipo"])
            estatus_descriptivo = ESTATUS_MAP.get(
                tipo_sueldo["estatus"], tipo_sueldo["estatus"]
            )
            ws.append(
                [
                    tipo_descriptivo,
                    float(
                        tipo_sueldo["monto"]
                    ),  # Convierte Decimal a float para evitar errores
                    tipo_sueldo["descripcion"] if tipo_sueldo["descripcion"] else "",
                    estatus_descriptivo,
                ]
            )

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="Tipos_Sueldos.xlsx")
