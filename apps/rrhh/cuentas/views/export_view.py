from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from rrhh.cuentas.models import Cuenta  # Asegúrate de importar el modelo Cuenta


class CuentaExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "rrhh.cuentas.exel_cuenta"  # Define el permiso necesario

    def get(self, request, *args, **kwargs):
        cuentas = Cuenta.objects.all().values(
            "banco",
            "tipo",
            "numero_cuenta",
            "pago_movil",
            "telefono",
            "empleado__nombres",  # Asume que el modelo Empleado tiene un campo 'nombres'
            "empleado__apellidos",  # Asume que el modelo Empleado tiene un campo 'apellidos'
        )

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:F1")  # Mezcla las celdas de A1 a F1
        ws["A1"] = "Listado de Cuentas"  # Agrega el texto del título
        ws["A1"].alignment = Alignment(
            horizontal="center"
        )  # Centra el texto del título
        ws["A1"].font = Font(
            bold=True, color="0000FF"
        )  # Establece el texto en negrita y color azul
        ws["A2"] = ""  # Agrega una fila vacía

        # Agrega los encabezados de las columnas
        ws.append(
            [
                "Banco",
                "Tipo de Cuenta",
                "Número de Cuenta",
                "Pago Móvil",
                "Teléfono",
                "Empleado",
            ]
        )

        # Ajusta el ancho de las columnas
        columnas = ws.column_dimensions
        columnas["A"].width = 30
        columnas["B"].width = 20
        columnas["C"].width = 20
        columnas["D"].width = 15
        columnas["E"].width = 15
        columnas["F"].width = 30

        # Agrega los datos de las cuentas
        for cuenta in cuentas:
            ws.append(
                [
                    cuenta["banco"],
                    cuenta["tipo"],
                    cuenta["numero_cuenta"],
                    "Sí" if cuenta["pago_movil"] else "No",
                    cuenta["telefono"],
                    f"{cuenta['empleado__nombres']} {cuenta['empleado__apellidos']}",
                ]
            )

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="Cuentas.xlsx")
