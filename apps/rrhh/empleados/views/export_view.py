from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from rrhh.empleados.models import Empleado  # Asegúrate de importar el modelo Empleado


class EmpleadoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "rrhh.empleados.exel_empleado"  # Define el permiso necesario

    def get(self, request, *args, **kwargs):
        empleados = Empleado.objects.all().values(
            "estatus",
            "nombres",
            "apellidos",
            "nacionalidad",
            "cedula",
            "sexo",
            "fecha_nacimiento",
            "estado_civil",
            "tipo_sangre",
            "email",
            "telefono",
            "direccion",
            "estudia",
            "discapacitado",
            "contratos",
        )

        # Crea un archivo Excel en memoria
        wb = Workbook()
        ws = wb.active

        # Agrega el título en la primera fila
        ws.merge_cells("A1:N1")  # Mezcla las celdas de A1 a N1
        ws["A1"] = "Listado de Empleados"  # Agrega el texto del título
        ws["A1"].alignment = Alignment(
            horizontal="center"
        )  # Centra el texto del título
        ws["A1"].font = Font(
            bold=True, color="0000FF"
        )  # Establece el texto en negrita y color azul
        ws["A2"] = ""  # Agrega una fila vacía

        # Agrega los encabezados de las columnas
        ws.append(
            [
                "Estatus",
                "Nombres",
                "Apellidos",
                "Nacionalidad",
                "Cédula",
                "Sexo",
                "Fecha de Nacimiento",
                "Estado Civil",
                "Tipo de Sangre",
                "Email",
                "Teléfono",
                "Dirección",
                "Estudia",
                "Discapacitado",
                "Contratos",
            ]
        )

        # Ajusta el ancho de las columnas
        columnas = ws.column_dimensions
        columnas["A"].width = 15
        columnas["B"].width = 20
        columnas["C"].width = 20
        columnas["D"].width = 15
        columnas["E"].width = 15
        columnas["F"].width = 10
        columnas["G"].width = 20
        columnas["H"].width = 15
        columnas["I"].width = 15
        columnas["J"].width = 25
        columnas["K"].width = 15
        columnas["L"].width = 30
        columnas["M"].width = 15
        columnas["N"].width = 15

        # Agrega los datos de los empleados
        for empleado in empleados:
            ws.append(
                [
                    empleado["estatus"],
                    empleado["nombres"],
                    empleado["apellidos"],
                    empleado["nacionalidad"],
                    empleado["cedula"],
                    empleado["sexo"],
                    (
                        empleado["fecha_nacimiento"].strftime("%Y-%m-%d")
                        if empleado["fecha_nacimiento"]
                        else ""
                    ),
                    empleado["estado_civil"],
                    empleado["tipo_sangre"],
                    empleado["email"],
                    empleado["telefono"],
                    empleado["direccion"],
                    "Sí" if empleado["estudia"] else "No",
                    "Sí" if empleado["discapacitado"] else "No",
                    empleado["contratos"],
                ]
            )

        # Convierte el archivo Excel en memoria a bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Devuelve el archivo Excel como respuesta
        return FileResponse(output, as_attachment=True, filename="Empleados.xlsx")
