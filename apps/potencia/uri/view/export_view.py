from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from potencia.uri.models import (
    Uri,
    ESTATUS_CHOICES2,
    ESTATUS_CHOICES3,
    ESTATUS_CHOICES4,
    ESTATUS_CHOICES5,
    ESTATUS_CHOICES6,
    ESTATUS_CHOICES7,
    ESTATUS_CHOICES8,
    ESTATUS_CHOICES9,
    ESTATUS_CHOICES10,
    ESTATUS_CHOICES11,
    ESTATUS_CHOICES12,
    ESTATUS_CHOICES13,
    ESTATUS_CHOICES14,
    ESTATUS_CHOICES15,
    ESTATUS_CHOICES16,
    ESTATUS_CHOICES17,
    ESTATUS_CHOICES18,
    ESTATUS_CHOICES19,
    GENERO_CHOICES,
    CONTACTO_CHOICES,
    NOCONTACTO_CHOICES,
    ACCIDENVEHI_CHOICES,
    ARMA_CHOICES,
    TRAUMAVE_CHOICES,
)


class URIExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = (
        "uri.exel_uri"  # Make sure this matches your model's permissions
    )

    def get_choice_display(self, value, choices):
        """Helper function to get display value from choices"""
        return dict(choices).get(value, value)

    def get(self, request, *args, **kwargs):
        # Fetch all URI records
        uris = Uri.objects.all()

        # Create an Excel file in memory
        wb = Workbook()
        ws = wb.active

        # Add title in the first row
        ws.merge_cells("A1:AZ1")  # Wide merge for title
        ws["A1"] = "Reporte de Unidades de Respuesta Inmediata (URI)"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF")
        ws.append([])  # Empty row after title

        # Headers - Organized by sections from the model
        headers = [
            # 1) Información General
            "ID",
            "Fecha de Atención",
            "Número de Reporte",
            "Placa",
            "Institución",
            "Tipo de Unidad",
            "Número Interno",
            # 2) Informacion legal
            "Contacto con el paciente",
            "No contacto con el paciente",
            "Servicio Asistencial",
            "Médico Receptor",
            "MS/DS",
            "Registro Fotográfico",
            # 2) Datos del Paciente
            "Nombre Paciente",
            "Apellido Paciente" "Cédula Paciente",
            "Teléfono Paciente",
            "Género Paciente",
            "Dirección Paciente",
            "Organismo Presente",
            "Jefe de Comisión",
            "Unidad/Placa Autoridad",
            "Firma Autoridad",
            # 3) Consentimiento Informado
            "Nombre Acompañante",
            "Apellido Acompañante",
            "Parentesco Acompañante",
            "Cédula Acompañante",
            "Teléfono Acompañante",
            "Género Acompañante",
            "Dirección Acompañante",
            "Nombre Testigo",
            "Apellido Testigo",
            "Edad Testigo",
            "Cédula Testigo",
            "Teléfono Testigo",
            "Dirección Testigo",
            # 4) Dirección del Evento
            "Estado",
            "Municipio",
            "Parroquia",
            "Sector/Urbanización",
            "Calle/Avenida/Carrera",
            "Edif/Casa",
            "Piso/Apto",
            "Punto de Referencia",
            "Lugar de Atención",
            "Modo de Traslado",
            "Vía del Reporte",
            "Tipo de Servicio",
            # Cronología del Servicio
            "Hora Alarma",
            "Hora Salida",
            "Hora Llegada",
            "Hospital Destino",
            "Transferencia Emergencia",
            "Hora Retorno Sede",
            "Tiempo de Servicio",
            "Observaciones Servicio",
            # 5) Información Clínica - Trauma
            "Accidente Vehicular",
            "Enfrentamiento Armado",
            "Trauma con Vehículo",
            "Viajaba como",
            "Sustancia Peligrosa",
            "Observaciones Sustancia",
            "Trauma no Intencional",
            "Emergencia Médica no Traumática",
            # Evaluación Inicial
            "Hemorragias",
            "Presión Directa",
            "Empaquetado Herida",
            "Torniquete",
            # Vía Aérea
            "Evaluación Vía Aérea",
            "Intervención Vía Aérea",
            "Resultado Vía Aérea",
            "Descripción Adicional Vía Aérea",
            # 6) Respiración, Oxigenación y Circulación
            "Evaluación Respiración",
            "Intervención Respiración",
            "Resultado Respiración",
            "Descripción Adicional Respiración",
            "Color Piel",
            "Temperatura Piel",
            "Humedad Piel",
            "Pulsos Distales",
            "Otras Heridas",
            "Fracturas",
            "Maniobra Pelvis",
            # Déficit Neurológico
            "ECG O",
            "ECG V",
            "ECG M",
            "ECG Total",
            "Reacción Pupilar",
            # Exposición
            "Hipotermia",
            "Signos/Síntomas",
            "Alergias",
            "Medicamentos",
            "Preexistencias",
            "Última Comida",
            "Evento",
            # 7) Signos Vitales
            "Hora Medición",
            "Frecuencia Cardiaca",
            "Frecuencia Respiratoria",
            "Presión Arterial",
            "SPO2",
            "Temperatura",
            "Llenado Capilar",
            "Glicemia Capilar",
            "Escala Glasgow",
            # Tratamiento
            "Medicamento",
            "Dosis",
            "Hora Tratamiento",
            "Resultado Evaluación",
            # Registro de Referencias
            "Traslado Inicial",
            "Hospital Origen",
            "Médico que Refiere",
            "Hora Salida Hospital",
            "Hospital Destino",
            "Hora Llegada Hospital",
            "Causa Referencia",
            # Personal Involucrado
            "Técnico Emergencias",
            "Cédula Técnico",
            "Tercer Tripulante",
            "Cédula Tripulante",
            "Conductor Unidad",
            "Cédula Conductor",
            "Supervisor Guardia",
            "Cédula Supervisor",
            "Médico Guardia",
            "Cédula Médico",
            "Sellos/MSDS",
        ]

        ws.append(headers)

        # Set column widths
        column_widths = {
            "A": 10,
            "B": 15,
            "C": 15,
            "D": 10,
            "E": 20,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 20,
            "J": 20,
            "K": 20,
            "L": 10,
            "M": 15,
            "N": 25,
            "O": 15,
            "P": 15,
            "Q": 15,
            "R": 25,
            "S": 20,
            "T": 15,
            "U": 20,
            "V": 15,
            "W": 15,
            "X": 15,
            "Y": 15,
            "Z": 15,
            "AA": 15,
            "AB": 15,
            "AC": 15,
            "AD": 15,
            "AE": 20,
            "AF": 20,
            "AG": 15,
            "AH": 15,
            "AI": 15,
            "AJ": 20,
            "AK": 15,
            "AL": 15,
            "AM": 15,
            "AN": 20,
            "AO": 20,
            "AP": 20,
            "AQ": 20,
            "AR": 20,
            "AS": 20,
            "AT": 20,
            "AU": 20,
            "AV": 20,
            "AW": 20,
            "AX": 20,
            "AY": 20,
            "AZ": 20,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Data rows
        for uri in uris:
            row = [
                # 1) Información General
                uri.id,
                uri.fecha_atencion,
                uri.nroreporte,
                uri.placa,
                uri.institucion,
                uri.tipounidad,
                uri.num_interna,
                self.get_choice_display(uri.contactopaciente, CONTACTO_CHOICES),
                self.get_choice_display(uri.contactonopaciente, NOCONTACTO_CHOICES),
                uri.ambulatorio,
                uri.servicioAsistencial,
                uri.medico_receptor,
                uri.msds,
                self.get_choice_display(uri.foto, ESTATUS_CHOICES2),
                # 2) Datos del Paciente
                uri.nombrepaciente,
                uri.apellidopaciente,
                uri.cedulapaciente,
                uri.telefonopaciente,
                self.get_choice_display(uri.generopaciente, GENERO_CHOICES),
                uri.direccionpaciente,
                uri.organismo,
                uri.jefedecomision,
                uri.unidad_placa,
                uri.firma,
                # 3) Consentimiento Informado
                uri.nombre_acompanante,
                uri.apellido_acompanante,
                uri.parentezco_acompanante,
                uri.cedula_acompanante,
                uri.telefono_acompanate,
                self.get_choice_display(uri.genero_acompanante, GENERO_CHOICES),
                uri.direccion_acompanante,
                uri.nombre_testigo,
                uri.apellido_testigo,
                uri.edad_testigo,
                uri.cedula_testigo,
                uri.telefono_testigo,
                uri.direccion_testigo,
                # 4) Dirección del Evento
                uri.estado,
                uri.municipio,
                uri.parroquia,
                uri.sector_evento,
                uri.calle_evento,
                uri.casa_evento,
                uri.piso_evento,
                uri.referencia_evento,
                self.get_choice_display(uri.lugar_atencion, ESTATUS_CHOICES3),
                self.get_choice_display(uri.modo_traslado, ESTATUS_CHOICES4),
                self.get_choice_display(uri.via_reporte, ESTATUS_CHOICES5),
                self.get_choice_display(uri.servicio_tipo, ESTATUS_CHOICES6),
                # Cronología del Servicio
                uri.hora_alarma,
                uri.hora_salida,
                uri.hora_llegada,
                uri.hospital,
                uri.transferencia_emergencia,
                uri.hora_sede,
                uri.tiempo_servicio,
                uri.observaciones_servicio,
                # 5) Información Clínica - Trauma
                self.get_choice_display(uri.accidenteVehicular, ACCIDENVEHI_CHOICES),
                self.get_choice_display(uri.enfrentamientoArmado, ARMA_CHOICES),
                self.get_choice_display(uri.traumaVehiculo, TRAUMAVE_CHOICES),
                self.get_choice_display(uri.viajaba, ESTATUS_CHOICES7),
                self.get_choice_display(uri.sustanciaPeligrosa, ESTATUS_CHOICES2),
                uri.observacionesSustancia,
                self.get_choice_display(uri.traumaNoIntencional, ESTATUS_CHOICES8),
                self.get_choice_display(uri.emergenciaMedica, ESTATUS_CHOICES9),
                # Evaluación Inicial
                self.get_choice_display(uri.hemorragia, ESTATUS_CHOICES2),
                self.get_choice_display(uri.presion, ESTATUS_CHOICES2),
                self.get_choice_display(uri.empaquetado, ESTATUS_CHOICES2),
                self.get_choice_display(uri.torniquete, ESTATUS_CHOICES2),
                # Vía Aérea
                self.get_choice_display(uri.evaluacion, ESTATUS_CHOICES10),
                self.get_choice_display(uri.intervencion, ESTATUS_CHOICES11),
                self.get_choice_display(uri.resultado, ESTATUS_CHOICES12),
                uri.descripcion_adic,
                # 6) Respiración, Oxigenación y Circulación
                self.get_choice_display(uri.evaluacionResp, ESTATUS_CHOICES13),
                self.get_choice_display(uri.intervencionResp, ESTATUS_CHOICES14),
                self.get_choice_display(uri.resultadoResp, ESTATUS_CHOICES15),
                uri.descripcion_adic_resp,
                self.get_choice_display(uri.colorPiel, ESTATUS_CHOICES16),
                self.get_choice_display(uri.temperaturaPiel, ESTATUS_CHOICES17),
                self.get_choice_display(uri.humedadPiel, ESTATUS_CHOICES18),
                self.get_choice_display(uri.pulso, ESTATUS_CHOICES2),
                self.get_choice_display(uri.otrasHerida, ESTATUS_CHOICES2),
                self.get_choice_display(uri.fractura, ESTATUS_CHOICES2),
                self.get_choice_display(uri.maniobraPelvis, ESTATUS_CHOICES2),
                # Déficit Neurológico
                uri.ecgO,
                uri.ecgV,
                uri.ecgM,
                uri.ecgTotal,
                self.get_choice_display(uri.reaccionPupilar, ESTATUS_CHOICES19),
                # Exposición
                self.get_choice_display(uri.hipotermia, ESTATUS_CHOICES2),
                uri.signosSintomas,
                uri.alergias,
                uri.medicamentos,
                uri.preexistencias,
                uri.ultimaComida,
                uri.evento,
                # 7) Signos Vitales
                uri.horaMedicion,
                uri.frecuenciaCardiaca,
                uri.frecuenciaRespiratoria,
                uri.presionArterial,
                uri.spo2,
                uri.temperatura,
                uri.llenadoCapilar,
                uri.glicemiaCapilar,
                uri.escalaGlasgow,
                # Tratamiento
                uri.medicamento,
                uri.dosis,
                uri.hora,
                uri.resultadoEvaluacion,
                # Registro de Referencias
                self.get_choice_display(uri.trasladoIncial, ESTATUS_CHOICES2),
                uri.hospitalOrigen,
                uri.medicoRefiere,
                uri.horaSalidaHosp,
                uri.hospitalDestino,
                uri.horaLlegadaHosp,
                uri.causa,
                # Personal Involucrado
                uri.tecnicoEmergencia,
                uri.cedulaTecnico,
                uri.tercerTripulante,
                uri.cedulaTripulante,
                uri.conductorUnidad,
                uri.cedulaConductor,
                uri.supervisorGuardia,
                uri.cedulaSupervisor,
                uri.medicoGuardia,
                uri.cedulaMedico,
                uri.selloMsds,
            ]

            ws.append(row)

        # Convert Excel file to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Return Excel file as response
        return FileResponse(output, as_attachment=True, filename="Reporte_URI.xlsx")
