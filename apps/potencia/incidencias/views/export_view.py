from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from potencia.incidencias.models import Incidencia
from helpers.BaseModelMixin import ESTADOS_CHOICES

ESTADOS_CHOICES = dict(ESTADOS_CHOICES)


class IncidenciaExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "potencia.listar_incidencia"

    def get(self, request, *args, **kwargs):
        # Obtener todas las incidencias con los campos necesarios
        incidencias = (
            Incidencia.objects.all()
            .select_related("sede", "departamento", "tipo_incidencia")
            .values(
                "sede__sede",
                "departamento__nombre",
                "tipo_incidencia__tipo",
                "estado",
                "tipo_solicitud",
                "observaciones",
                "created_at",
            )
        )

        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active

        # Configurar el título del reporte
        ws.merge_cells("A1:G1")
        ws["A1"] = "Reporte de Incidencias"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF")
        ws.append([])  # Espacio en blanco

        # Encabezados de columnas
        ws.append(
            [
                "Sede",
                "Departamento",
                "Tipo de Incidencia",
                "Estado",
                "Tipo de Solicitud",
                "Observaciones",
                "Fecha Creación",
            ]
        )

        # Configurar anchos de columnas
        column_widths = {
            "A": 20,  # Sede
            "B": 20,  # Departamento
            "C": 20,  # Tipo de Incidencia
            "D": 15,  # Estado
            "E": 20,  # Tipo de Solicitud
            "F": 40,  # Observaciones
            "G": 15,  # Fecha Creación
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Llenar con datos
        for incidencia in incidencias:
            estado_codigo = incidencia["estado"]
            estado_nombre = ESTADOS_CHOICES.get(estado_codigo, estado_codigo)

            ws.append(
                [
                    incidencia["sede__sede"],
                    incidencia["departamento__nombre"],
                    incidencia["tipo_incidencia__tipo"],
                    estado_nombre,
                    incidencia["tipo_solicitud"],
                    incidencia["observaciones"],
                    (
                        incidencia["created_at"].strftime("%Y-%m-%d")
                        if incidencia["created_at"]
                        else ""
                    ),
                ]
            )

        # Preparar la respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="Incidencias.xlsx",
        )
