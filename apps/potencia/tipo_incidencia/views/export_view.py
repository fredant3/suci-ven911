from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from potencia.tipo_incidencia.models import TipoIncidencia


class TipoIncidenciaExcelApiView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "tipo_incidencia.listar_tipo_incidencia"

    def get(self, request, *args, **kwargs):
        # Consulta todos los tipos de incidencia
        tipos_incidencia = TipoIncidencia.objects.filter(deleted_at__isnull=True).values(
            "id",
            "tipo",
            "created_at",
            "updated_at",
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Tipos de Incidencia"

        # Configuración del título
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "REGISTRO DE TIPOS DE INCIDENCIA"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados
        headers = [
            ("ID", 10),
            ("Tipo de Incidencia", 40),
            ("Fecha Creación", 20),
            ("Fecha Actualización", 20),
        ]

        # Configurar columnas
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Llenar datos
        for row_num, tipo in enumerate(tipos_incidencia, 3):
            ws.append(
                [
                    tipo["id"],
                    tipo["tipo"],
                    tipo["created_at"].strftime("%d/%m/%Y %H:%M") if tipo["created_at"] else "N/A",
                    tipo["updated_at"].strftime("%d/%m/%Y %H:%M") if tipo["updated_at"] else "N/A",
                ]
            )

        # Aplicar bordes
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output, as_attachment=True, filename="reporte_tipos_incidencia.xlsx"
        )