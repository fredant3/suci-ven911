from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from administracion.tipo_averia.models import TipoAveria


class TipoAveriaExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "administracion.tipo_averia.listar_tipo_averia"

    def get(self, request, *args, **kwargs):
        tipos = TipoAveria.objects.all().values(
            "nombre",
            "created_at",
        )

        wb = Workbook()
        ws = wb.active

        # Configuración del título
        ws.merge_cells("A1:B1")
        ws["A1"] = "Reporte de Tipos de Avería"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF")
        ws.append([])

        # Encabezados
        ws.append(["Nombre", "Fecha de Creación"])

        # Ajustes de columnas
        column_widths = {"A": 40, "B": 20}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Datos
        for tipo in tipos:
            fecha = (
                tipo["created_at"].strftime("%Y-%m-%d") if tipo["created_at"] else ""
            )
            ws.append([tipo["nombre"], fecha])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="tipos_averia.xlsx",
        )
