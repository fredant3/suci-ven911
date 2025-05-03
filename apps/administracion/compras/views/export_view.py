from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from administracion.compras.model import Compra  # Importar el modelo Compra


class CompraExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = (
        "administracion.listar_compra"  # Asegúrate de usar el nombre correcto de tu app
    )

    def get(self, request, *args, **kwargs):
        # Obtener datos con select_related para optimización
        compras = (
            Compra.objects.select_related("articulo")
            .all()
            .values(
                "articulo__nombre",
                "n_orden",
                "valor_bs",
                "created_at",  # Usamos el campo created_at de BaseModel
            )
        )

        # Crear libro Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Compras"

        # Configuración del título
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "REGISTRO DE COMPRAS"
        title_cell.font = Font(bold=True, size=14, color="0047AB")  # Azul corporativo
        title_cell.alignment = Alignment(horizontal="center")

        # Encabezados con estilo
        headers = [
            ("Artículo", 30),
            ("N° Orden", 15),
            ("Valor (BS)", 20),
            ("Fecha Registro", 20),
        ]

        ws.append([header[0] for header in headers])  # Añadir encabezados

        # Aplicar estilos a los encabezados
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=2, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[chr(64 + col)].width = headers[col - 1][1]

        # Llenar datos
        for compra in compras:
            fecha = (
                compra["created_at"].strftime("%d/%m/%Y")
                if compra["created_at"]
                else "N/A"
            )
            ws.append(
                [
                    compra["articulo__descripcion"],  # Descripción del artículo
                    f"ORD-{compra['n_orden']}",  # Formato personalizado
                    f"Bs. {compra['valor_bs']}",  # Formato monetario
                    fecha,
                ]
            )

        # Ajustar alineación de columnas
        for row in ws.iter_rows(min_row=3, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal="left")

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="Compras.xlsx",  # Nombre del archivo en minúsculas
        )
