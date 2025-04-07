from io import BytesIO
from administracion.asignaciones.models import Asignacion  # Cambio de modelo
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


class AsignacionExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "administracion.listar_asignacion"

    def get(self, request, *args, **kwargs):
        asignaciones = Asignacion.objects.all().values(
            "articulo__serial",
            "sede__sede",
            "departamento__nombre",
            "cantidad",
            "descripcion",
            "observaciones",
            "created_at",
        )

        wb = Workbook()
        ws = wb.active

        ws.merge_cells("A1:G1")
        ws["A1"] = "Reporte de asignaciones"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF")
        ws.append([])

        ws.append(
            [
                "Artículo",
                "Sede",
                "Departamento",
                "Cantidad",
                "Descripción",
                "Observaciones",
                "Fecha de creación",
            ]
        )

        column_widths = {
            "A": 25,
            "B": 20,
            "C": 25,
            "D": 12,
            "E": 40,
            "F": 40,
            "G": 20,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for registro in asignaciones:
            ws.append(
                [
                    registro["articulo__serial"],
                    registro["sede__sede"],
                    registro["departamento__nombre"],
                    registro["cantidad"],
                    registro["descripcion"],
                    registro["observaciones"],
                    (
                        registro["created_at"].strftime("%Y-%m-%d")
                        if registro["created_at"]
                        else ""
                    ),
                ]
            )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="Asignaciones.xlsx",
        )
