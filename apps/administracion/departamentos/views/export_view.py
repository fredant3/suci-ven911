from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from administracion.departamentos.models import (
    Departamento,
)  # Importar el modelo Departamento


class DepartamentoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = (
        "administracion.listar_departamento"  # Usar el permiso correspondiente
    )

    def get(self, request, *args, **kwargs):
        departamentos = Departamento.objects.all().values("nombre", "created_at")

        wb = Workbook()
        ws = wb.active

        # Configuración del título
        ws.merge_cells("A1:B1")
        ws["A1"] = "Reporte de Departamentos"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF")
        ws.append([])  # Espacio en blanco

        # Encabezados
        ws.append(["Nombre del Departamento", "Fecha de Creación"])

        # Ajustes de columnas
        column_widths = {"A": 30, "B": 20}  # Nombre  # Fecha
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Datos
        for departamento in departamentos:
            fecha = (
                departamento["created_at"].strftime("%Y-%m-%d")
                if departamento["created_at"]
                else ""
            )
            ws.append([departamento["nombre"], fecha])

        # Generar respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="Departamentos.xlsx",  # Nombre del archivo en minúsculas
        )
