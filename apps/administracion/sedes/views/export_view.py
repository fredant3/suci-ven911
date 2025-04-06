from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from administracion.sedes.models import (
    Sede,
    ESTATUS_CHOICES,
)  # Importar el modelo y choices


class SedeExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "tu_app.listar_sede"  # Ajustar con tu app real

    def get(self, request, *args, **kwargs):
        # Consulta optimizada con select_related si hay relaciones
        sedes = Sede.objects.all().values("sede", "estatus", "created_at")

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Sedes"

        # Configuración del título
        ws.merge_cells("A1:C1")
        title_cell = ws["A1"]
        title_cell.value = "REPORTE DE SEDES"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados
        headers = [("Nombre Sede", 30), ("Estatus", 20), ("Fecha Registro", 20)]

        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Mapeo de estatus
        estatus_map = dict(ESTATUS_CHOICES)

        # Datos
        for row_num, sede in enumerate(sedes, 3):
            ws.append(
                [
                    sede["sede"],
                    estatus_map.get(
                        sede["estatus"], "Desconocido"
                    ),  # Mapear valor choice
                    (
                        sede["created_at"].strftime("%d/%m/%Y")
                        if sede["created_at"]
                        else "N/A"
                    ),
                ]
            )

        # Aplicar bordes
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(output, as_attachment=True, filename="reporte_sedes.xlsx")
