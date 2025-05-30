from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from administracion.averia.models import Averia


class AveriaExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "administracion.listar_averia"

    def get(self, request, *args, **kwargs):
        averias = Averia.objects.all().values(
            "tipo_averia__nombre",
            "departamento__nombre",
            "problema",
            "serial",
            "codigo_bn",
            "created_at",
            "d_averia",
            "observaciones",  # Campo añadido en lugar de ubicacion
        )

        wb = Workbook()
        ws = wb.active

        # Configuración del título
        ws.merge_cells("A1:H1")
        ws["A1"] = "Reporte de Averías"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF")
        ws.append([])  # Espacio en blanco

        # Encabezados
        ws.append(
            [
                "Tipo de Avería",
                "Departamento",
                "Problema",
                "Serial",
                "Código BN",
                "Departamento que reporta",
                "Observaciones",  # Reemplazado Ubicación por Observaciones
                "Fecha de Creación",
            ]
        )

        # Ajustes de columnas
        column_widths = {
            "A": 25,  # Tipo de Avería
            "B": 25,  # Departamento
            "C": 40,  # Problema
            "D": 20,  # Serial
            "E": 20,  # Código BN
            "F": 30,  # Departamento que reporta
            "G": 40,  # Observaciones
            "H": 20,  # Fecha
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Mapeo de DEP_AVERIA
        DEP_AVERIA_MAP = {
            "1": "Asesoría Jurídica",
            "2": "Gestión Humana",
            "3": "Gestión Administrativa",
            "4": "Unidad de Respuesta Inmediata",
            "5": "Potencia",
            "6": "Organización",
            "7": "Presupuestos",
            "8": "Planificación",
            "9": "Protección y Seguridad Integral",
            "10": "Biblioteca de Manuales",
            "11": "Operaciones",
            "12": "Tecnología Comunicación e Información",
            "13": "Gestion Comunicacional",
        }

        # Datos
        for averia in averias:
            fecha = (
                averia["created_at"].strftime("%Y-%m-%d")
                if averia["created_at"]
                else ""
            )
            ws.append(
                [
                    averia["tipo_averia__nombre"],
                    averia["departamento__nombre"],
                    averia["problema"],
                    averia["serial"],
                    averia["codigo_bn"],
                    DEP_AVERIA_MAP.get(averia["d_averia"], "Desconocido"),
                    averia["observaciones"] or "",  # Manejo de valores nulos
                    fecha,
                ]
            )

        # Generar respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="averias.xlsx",
        )
