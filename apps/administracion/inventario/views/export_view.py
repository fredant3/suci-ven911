from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from administracion.inventario.models import Articulo


class ArticuloExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "administracion.listar_articulo"

    def get(self, request, *args, **kwargs):
        articulos = Articulo.objects.select_related("tipo_articulo").values(
            "descripcion",
            "marca",
            "modelo",
            "serial",
            "placa",
            "cantidad",
            "tipo_articulo__nombre",
            "fecha_adq",
            "asignado",
            "codigo_bn",
            "cantidad_combustible",
            "created_at",
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Artículos"

        # Configuración del título
        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = "INVENTARIO DE ARTÍCULOS"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados (ahora con 11 columnas)
        headers = [
            ("Descripción", 30),
            ("Marca", 20),
            ("Modelo", 20),
            ("Serial", 20),
            ("Placa", 15),
            ("Cantidad", 10),
            ("Tipo Artículo", 20),
            ("Fecha Adquisición", 15),
            ("Asignado", 10),
            ("Código BN", 15),
            ("Combustible (L)", 15),
            ("Fecha Registro", 20),
        ]

        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Datos (sin el campo condicion)
        for row_num, articulo in enumerate(articulos, 3):
            ws.append(
                [
                    articulo["descripcion"],
                    articulo["marca"] or "-",
                    articulo["modelo"] or "-",
                    articulo["serial"] or "-",
                    articulo["placa"] or "-",
                    articulo["cantidad"],
                    articulo["tipo_articulo__nombre"],
                    (
                        articulo["fecha_adq"].strftime("%d/%m/%Y")
                        if articulo["fecha_adq"]
                        else "-"
                    ),
                    "Sí" if articulo["asignado"] == "yes" else "No",
                    articulo["codigo_bn"] or "-",
                    articulo["cantidad_combustible"] or "-",
                    (
                        articulo["created_at"].strftime("%d/%m/%Y")
                        if articulo["created_at"]
                        else "-"
                    ),
                ]
            )

        # Aplicar bordes
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output, as_attachment=True, filename="inventario_articulos.xlsx"
        )
