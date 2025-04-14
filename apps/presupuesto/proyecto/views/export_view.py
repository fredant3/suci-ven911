from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from presupuesto.proyecto.models import Proyecto


class ProyectoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "proyecto.listar_proyecto"

    def get(self, request, *args, **kwargs):
        # Consulta optimizada con todos los campos relevantes
        proyectos = Proyecto.objects.filter(deleted_at__isnull=True).values(
            "nombrep",
            "fechai",
            "fechac",
            "situacionp",
            "montoproyecto",
            "responsableg",
            "responsablet",
            "responsabler",
            "responsablea",
            "estatus",
            "created_at"
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Registro de Proyectos"

        # Configuración del título (ajustado a 11 columnas)
        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = "REGISTRO DE PROYECTOS"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados con anchos personalizados
        headers = [
            ("Nombre Proyecto", 30),
            ("Fecha Inicio", 15),
            ("Fecha Culminación", 15),
            ("Situación Presup.", 25),
            ("Monto Total", 20),
            ("Resp. Gerente", 25),
            ("Resp. Técnico", 25),
            ("Resp. Registrador", 25),
            ("Resp. Administrativo", 25),
            ("Estatus", 20),
            ("Fecha Registro", 20)
        ]

        # Configurar columnas
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num) if col_num <= 26 else 'A' + chr(64 + col_num - 26)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Llenar datos
        for row_num, proyecto in enumerate(proyectos, 3):
            ws.append([
                proyecto["nombrep"],
                proyecto["fechai"].strftime("%d/%m/%Y") if proyecto["fechai"] else "N/A",
                proyecto["fechac"].strftime("%d/%m/%Y") if proyecto["fechac"] else "N/A",
                proyecto["situacionp"],
                proyecto["montoproyecto"],
                proyecto["responsableg"],
                proyecto["responsablet"],
                proyecto["responsabler"],
                proyecto["responsablea"],
                proyecto["estatus"],
                proyecto["created_at"].strftime("%d/%m/%Y %H:%M") if proyecto["created_at"] else "N/A"
            ])

        # Aplicar bordes a todas las celdas
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="reporte_proyectos.xlsx"
        )