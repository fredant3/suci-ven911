from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from presupuesto.cedente.models import Cedente


class CedenteExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "cedente.listar_cedente"

    def get(self, request, *args, **kwargs):
        # Query with all financial fields
        cedentes = Cedente.objects.filter(deleted_at__isnull=True).values(
            "idc",
            "partidac",
            "generalc",
            "espefc",
            "subespefc",
            "denomc",
            "presuacorc",
            "caufechac",
            "dispc",
            "montocc",
            "saldofc",
            "direccionc",
            "created_at"
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Registro de Cedentes"

        # Title configuration (13 columns)
        ws.merge_cells("A1:M1")
        title_cell = ws["A1"]
        title_cell.value = "REGISTRO DE CEDENTES PRESUPUESTARIOS"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Financial headers with optimized widths
        headers = [
            ("ID Cedente", 15),
            ("Partida Contable", 20),
            ("General", 15),
            ("Especificación", 20),
            ("Sub-Especialidad", 20),
            ("Denominación", 25),
            ("Presupuesto Asignado", 20),
            ("Causado a Fecha", 20),
            ("Disponible", 15),
            ("Monto Comprometido", 20),
            ("Saldo Final", 15),
            ("Dirección Cedente", 25),
            ("Fecha Registro", 20)
        ]

        # Configure columns with financial emphasis
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num) if col_num <= 26 else 'A' + chr(64 + col_num - 26)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            # Highlight financial headers
            if header in ["Presupuesto Asignado", "Causado a Fecha", "Disponible", 
                         "Monto Comprometido", "Saldo Final"]:
                cell.font = Font(bold=True, color="0047AB")

        # Fill financial data
        for row_num, cedente in enumerate(cedentes, 3):
            ws.append([
                cedente["idc"],
                cedente["partidac"],
                cedente["generalc"],
                cedente["espefc"],
                cedente["subespefc"],
                cedente["denomc"],
                cedente["presuacorc"],
                cedente["caufechac"],
                cedente["dispc"],
                cedente["montocc"],
                cedente["saldofc"],
                cedente["direccionc"],
                cedente["created_at"].strftime("%d/%m/%Y %H:%M") if cedente["created_at"] else "N/A"
            ])

        # Apply professional borders
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generate financial report
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="reporte_cedentes.xlsx"
        )
