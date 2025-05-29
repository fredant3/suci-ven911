# asignaciones/views/export_view.py
from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from django.db.models import Q
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import date
from presupuesto.asignacion.models import Asignacion


class AsignacionExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "asignaciones.listar_asignar_presupuesto"

    def get(self, request, *args, **kwargs):
        # Obtener término de búsqueda
        search_term = request.GET.get("search", "")

        # Obtener asignaciones con relaciones de partida
        asignaciones = Asignacion.objects.filter(
            deleted_at__isnull=True
        ).select_related("partida")

        # Aplicar filtro de búsqueda
        if search_term:
            asignaciones = asignaciones.filter(
                Q(departamento__icontains=search_term)
                | Q(presupuesto__icontains=search_term)
                | Q(objetivo__icontains=search_term)
                | Q(partida__codigo__icontains=search_term)
                | Q(partida__titulo__icontains=search_term)
            )

        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Asignaciones Presupuestarias"

        # Configurar título
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"REPORTE DE ASIGNACIONES PRESUPUESTARIAS - {date.today().strftime('%d/%m/%Y')}"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados
        headers = [
            ("ID", 10),
            ("Partida", 25),
            ("Departamento", 30),
            ("Presupuesto", 20),
            ("Objetivo", 40),
            ("Fecha Creación", 15),
        ]

        # Configurar columnas y encabezados
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = (
                chr(64 + col_num) if col_num <= 26 else "A" + chr(64 + col_num - 26)
            )
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            if header in ["Presupuesto", "Código Partida"]:
                cell.font = Font(bold=True, color="0047AB")

        # Llenar datos
        for asignacion in asignaciones:
            partida_info = (
                f"{asignacion.partida.codigo} - {asignacion.partida.titulo}"
                if asignacion.partida
                else "N/A"
            )

            ws.append(
                [
                    asignacion.id,
                    partida_info,
                    asignacion.departamento,
                    asignacion.presupuesto,
                    asignacion.objetivo,
                    (
                        asignacion.created_at.strftime("%d/%m/%Y")
                        if asignacion.created_at
                        else "N/A"
                    ),
                ]
            )

        # Aplicar bordes
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename=f"Reporte_Asignaciones_{date.today().strftime('%Y%m%d')}.xlsx",
        )
