from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from django.db.models import Q
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from datetime import date
from presupuesto.partida.models import Partida


class AccionExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "acciones.listar_accion"

    def get(self, request, *args, **kwargs):
        # Get search parameter from request
        search_term = request.GET.get("search", "")

        # Base queryset
        partidas = Partida.objects.all()

        # Apply search filter if term exists
        if search_term:
            partidas = partidas.filter(
                Q(proyecto__icontains=search_term)
                | Q(situacion_presupuestaria__icontains=search_term)
                | Q(monto__icontains=search_term)
                | Q(responsable_gerente__icontains=search_term)
                | Q(responsable_tecnico__icontains=search_term)
                | Q(responsable_registrador__icontains=search_term)
                | Q(responsable_administrativo__icontains=search_term)
                | Q(estatus__icontains=search_term)
            )

        # Prepare data for export
        data = partidas.values(
            "codigo",
            "titulo",
        )

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Partidas"

        # Title and headers
        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = f"Reporte de Partidas - {date.today().strftime('%Y-%m-%d')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        headers = [
            "codigo",
            "titulo",
        ]
        ws.append(headers)

        # Set column widths
        column_widths = {
            "A": 30,  # codigo
            "B": 15,  # titulo
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Format header row
        for cell in ws[2]:  # Row 2 contains headers
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for accion in data:
            ws.append(
                [
                    Partida["partida"],
                    Partida["codigo"],
                    Partida["partida"],
                    (
                        Partida["created_at"].strftime("%Y-%m-%d")
                        if Partida["created_at"]
                        else ""
                    ),
                ]
            )

        # Prepare response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename=f"Acciones_{date.today().strftime('%Y%m%d')}.xlsx",
        )
