from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from django.db.models import Q
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from datetime import date
from presupuesto.acciones.models import Accion


class AccionExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "acciones.listar_accion"

    def get(self, request, *args, **kwargs):
        # Get search parameter from request
        search_term = request.GET.get("search", "")

        # Base queryset
        acciones = Accion.objects.all()

        # Apply search filter if term exists
        if search_term:
            acciones = acciones.filter(
                Q(proyecto__icontains=search_term)
                | Q(situacion_presupuestaria__icontains=search_term)
                | Q(monto__icontains=search_term)
                | Q(responsable_gerente__icontains=search_term)
                | Q(responsable_tecnico__icontains=search_term)
                | Q(responsable_registrador__icontains=search_term)
                | Q(responsable_administrativo__icontains=search_term)
                | Q(estatus__icontains=search_term)
            )

        # Prepare data for export
        data = acciones.values(
            "proyecto",
            "fecha_inicio",
            "fecha_culminacion",
            "situacion_presupuestaria",
            "monto",
            "responsable_gerente",
            "responsable_tecnico",
            "responsable_registrador",
            "responsable_administrativo",
            "created_at",
        )

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Acciones"

        # Title and headers
        ws.merge_cells("A1:K1")
        title_cell = ws["A1"]
        title_cell.value = f"Reporte de Acciones - {date.today().strftime('%Y-%m-%d')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")

        headers = [
            "Proyecto",
            "Fecha Inicio",
            "Fecha Culminación",
            "Situación Presupuestaria",
            "Monto",
            "Responsable Gerente",
            "Responsable Técnico",
            "Responsable Registrador",
            "Responsable Administrativo",
            "Estatus",
            "Fecha Creación",
        ]
        ws.append(headers)

        # Set column widths
        column_widths = {
            "A": 30,  # Proyecto
            "B": 15,  # Fecha Inicio
            "C": 18,  # Fecha Culminación
            "D": 25,  # Situación Presupuestaria
            "E": 15,  # Monto
            "F": 25,  # Responsable Gerente
            "G": 25,  # Responsable Técnico
            "H": 25,  # Responsable Registrador
            "I": 25,  # Responsable Administrativo
            "K": 15,  # Fecha Creación
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Format header row
        for cell in ws[2]:  # Row 2 contains headers
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Add data rows
        for accion in data:
            ws.append(
                [
                    accion["proyecto"],
                    (
                        accion["fecha_inicio"].strftime("%Y-%m-%d")
                        if accion["fecha_inicio"]
                        else ""
                    ),
                    (
                        accion["fecha_culminacion"].strftime("%Y-%m-%d")
                        if accion["fecha_culminacion"]
                        else ""
                    ),
                    accion["situacion_presupuestaria"],
                    accion["monto"],
                    accion["responsable_gerente"],
                    accion["responsable_tecnico"],
                    accion["responsable_registrador"],
                    accion["responsable_administrativo"],
                    (
                        accion["created_at"].strftime("%Y-%m-%d")
                        if accion["created_at"]
                        else ""
                    ),
                ]
            )

        # Prepare response
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename=f"Acciones_{date.today().strftime('%Y%m%d')}.xlsx",
        )
