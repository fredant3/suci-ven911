from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from presupuesto.cedente.models import Cedente
from presupuesto.receptor.models import (
    Receptor,
)


class TraspasoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "cedente.listar_cedente"

    def get(self, request, *args, **kwargs):
        # Obtener datos de cedentes
        cedentes = Cedente.objects.filter(deleted_at__isnull=True).values(
            "idc",
            "partidac",
            "generalc",
            "espefc",
            "subespefc",
            "denomc",
            "presuacorc",
            "caufechac",
            "dispc",
            "montocc",
            "saldofc",
            "direccionc",
            "created_at",
        )

        # Obtener datos de receptores
        receptores = Receptor.objects.filter(deleted_at__isnull=True).values(
            "idr",
            "partidar",
            "generalr",
            "espefr",
            "subespefr",
            "denomr",
            "presuacorr",
            "caufechar",
            "dispr",
            "montocr",
            "saldofr",
            "direccionr",
            "created_at",
            "cedente__idc",  # Incluir referencia al cedente
        )

        wb = Workbook()

        # Hoja para Cedentes
        ws_cedentes = wb.active
        ws_cedentes.title = "Registro de Cedentes"
        self._create_sheet(
            ws_cedentes,
            "REGISTRO DE CEDENTES PRESUPUESTARIOS",
            cedentes,
            [
                ("ID Cedente", 15),
                ("Partida Contable", 20),
                ("General", 15),
                ("Especificación", 20),
                ("Sub-Especialidad", 20),
                ("Denominación", 25),
                ("Presupuesto Asignado", 20),
                ("Causado a Fecha", 20),
                ("Disponible", 15),
                ("Monto Comprometido", 20),
                ("Saldo Final", 15),
                ("Dirección Cedente", 25),
                ("Fecha Registro", 20),
            ],
        )

        # Hoja para Receptores
        ws_receptores = wb.create_sheet(title="Registro de Receptores")
        self._create_sheet(
            ws_receptores,
            "REGISTRO DE RECEPTORES PRESUPUESTARIOS",
            receptores,
            [
                ("ID Receptor", 15),
                ("Partida Contable", 20),
                ("General", 15),
                ("Especificación", 20),
                ("Sub-Especialidad", 20),
                ("Denominación", 25),
                ("Presupuesto Acordado", 20),
                ("Causado a Fecha", 20),
                ("Disponible", 15),
                ("Monto a Ceder", 20),
                ("Saldo Final", 15),
                ("Dirección Receptor", 25),
                ("ID Cedente", 15),
                ("Fecha Registro", 20),
            ],
        )

        # Generar reporte
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output, as_attachment=True, filename="reporte_cedentes_y_receptores.xlsx"
        )

    def _create_sheet(self, worksheet, title, data, headers):
        """Método auxiliar para crear hojas con formato similar"""
        # Configurar título
        worksheet.merge_cells("A1:M1" if len(headers) <= 13 else "A1:N1")
        title_cell = worksheet["A1"]
        title_cell.value = title
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Configurar columnas
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = (
                chr(64 + col_num) if col_num <= 26 else "A" + chr(64 + col_num - 26)
            )
            worksheet.column_dimensions[col_letter].width = width
            cell = worksheet.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            if (
                "Presupuesto" in header
                or "Causado" in header
                or "Disponible" in header
                or "Monto" in header
                or "Saldo" in header
            ):
                cell.font = Font(bold=True, color="0047AB")

        # Llenar datos
        for row_num, item in enumerate(data, 3):
            row_values = []
            for field in item.keys():
                value = item[field]
                if field == "created_at" and value:
                    value = value.strftime("%d/%m/%Y %H:%M")
                row_values.append(value)
            worksheet.append(row_values)

        # Aplicar bordes
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in worksheet.iter_rows(
            min_row=2, max_row=worksheet.max_row, max_col=len(headers)
        ):
            for cell in row:
                cell.border = border
