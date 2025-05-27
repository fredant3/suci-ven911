from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from presupuesto.cedente.models import Cedente


class TraspasoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "cedente.listar_cedente"

    def get(self, request, *args, **kwargs):
        # Obtener datos de cedentes y sus receptores relacionados
        cedentes = (
            Cedente.objects.filter(deleted_at__isnull=True)
            .select_related("partida")
            .prefetch_related("receptor_set")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Traspasos Presupuestarios"

        # Configurar título
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = (
            "REGISTRO DE TRASPASOS PRESUPUESTARIOS (CEDENTES Y RECEPTORES)"
        )
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados combinados
        headers = [
            # Cedente
            ("ID", 10),
            ("Partida C.", 25),
            ("Denominación C.", 30),
            ("Presupuesto C.", 15),
            ("Causado C.", 15),
            ("Disponible C.", 15),
            ("Monto C.", 15),
            ("Saldo C.", 15),
            # Receptor
            ("ID Receptor", 10),
            ("Partida R.", 25),
            ("Denominación R.", 30),
            ("Monto a Ceder", 15),
            ("Saldo R.", 15),
            ("Fecha Traspaso", 20),
        ]

        # Configurar columnas
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = (
                chr(64 + col_num) if col_num <= 26 else "A" + chr(64 + col_num - 26)
            )
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            if any(
                x in header
                for x in ["Presupuesto", "Causado", "Disponible", "Monto", "Saldo"]
            ):
                cell.font = Font(bold=True, color="0047AB")

        # Llenar datos combinados
        row_num = 3
        for cedente in cedentes:
            # Datos del cedente
            partida_cedente = (
                f"{cedente.partida.codigo} - {cedente.partida.titulo}"
                if cedente.partida
                else "N/A"
            )

            cedente_data = [
                cedente.id,
                partida_cedente,
                cedente.denomc,
                cedente.presuacorc,
                cedente.caufechac,
                cedente.dispc,
                cedente.montocc,
                cedente.saldofc,
            ]

            # Obtener receptores relacionados
            receptores = cedente.receptor_set.all()

            if receptores.exists():
                for receptor in receptores:
                    partida_receptor = (
                        f"{receptor.partida.codigo} - {receptor.partida.titulo}"
                        if receptor.partida
                        else "N/A"
                    )

                    # Combinar datos del cedente con cada receptor
                    row_data = cedente_data + [
                        receptor.id,
                        partida_receptor,
                        receptor.denomr,
                        receptor.montocr,
                        receptor.saldofr,
                        (
                            receptor.created_at.strftime("%d/%m/%Y")
                            if receptor.created_at
                            else "N/A"
                        ),
                    ]
                    ws.append(row_data)
            else:
                # Si no hay receptores, mostrar solo datos del cedente
                row_data = cedente_data + ["", "", "", "", ""]
                ws.append(row_data)

            row_num += 1

        # Aplicar bordes
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar reporte
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output, as_attachment=True, filename="reporte_traspasos_combinados.xlsx"
        )
