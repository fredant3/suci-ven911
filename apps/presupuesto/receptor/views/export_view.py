from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from presupuesto.receptor.models import Receptor


class ReceptorExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "receptor.listar_receptor"

    def get(self, request, *args, **kwargs):
        # Query with all relevant fields
        receptores = Receptor.objects.filter(deleted_at__isnull=True).values(
            "idr",
            "partidar",
            "generalr",
            "espefr",
            "subespefr",
            "denomr",
            "presuacorr",
            "caufechar",
            "dispr",
            "montocr",
            "saldofr",
            "direccionr",
            "created_at"
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Registro de Receptores"

        # Title configuration (13 columns)
        ws.merge_cells("A1:M1")
        title_cell = ws["A1"]
        title_cell.value = "REGISTRO DE RECEPTORES PRESUPUESTARIOS"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Headers with custom widths
        headers = [
            ("ID Receptor", 15),
            ("Partida Contable", 20),
            ("General", 15),
            ("Especificación", 20),
            ("Sub-Especialidad", 20),
            ("Denominación", 25),
            ("Presupuesto Acordado", 20),
            ("Causado a Fecha", 20),
            ("Disponible", 15),
            ("Monto a Ceder", 15),
            ("Saldo Final", 15),
            ("Dirección Cedente", 25),
            ("Fecha Registro", 20)
        ]

        # Configure columns
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num) if col_num <= 26 else 'A' + chr(64 + col_num - 26)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Fill data
        for row_num, receptor in enumerate(receptores, 3):
            ws.append([
                receptor["idr"],
                receptor["partidar"],
                receptor["generalr"],
                receptor["espefr"],
                receptor["subespefr"],
                receptor["denomr"],
                receptor["presuacorr"],
                receptor["caufechar"],
                receptor["dispr"],
                receptor["montocr"],
                receptor["saldofr"],
                receptor["direccionr"],
                receptor["created_at"].strftime("%d/%m/%Y %H:%M") if receptor["created_at"] else "N/A"
            ])

        # Apply borders
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generate file
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="reporte_receptores.xlsx"
        )