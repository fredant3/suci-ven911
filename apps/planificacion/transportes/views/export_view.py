from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from planificacion.transportes.models import Transporte
from helpers.BaseModelMixin import ESTADOS_CHOICES, MONTH_CHOICES


ESTADOS_DISPLAY = dict(ESTADOS_CHOICES)
MESES_DISPLAY = dict(MONTH_CHOICES)


class TransporteExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "transportes.listar_transporte"

    def get(self, request, *args, **kwargs):
        transportes = Transporte.objects.all().values(
            "estado",
            "mes",
            "transporte",
            "cantidad",
            "created_at",
        )

        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active

        # Configurar el título del reporte
        ws.merge_cells("A1:G1")
        ws["A1"] = "Reporte de Transportes"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF", size=14)
        ws.append([])  # Espacio en blanco

        # Encabezados de columnas con estilo
        headers = [
            "Estado",
            "Mes",
            "Tipo de Transporte",
            "Cantidad",
            "Fecha Creación",
        ]
        ws.append(headers)

        # Estilo para los encabezados
        for cell in ws[3]:  # Fila 3 contiene los encabezados
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Configurar anchos de columnas
        column_widths = {
            "A": 20,  # Estado
            "B": 15,  # Mes
            "C": 30,  # Tipo de Transporte
            "D": 15,  # Cantidad
            "E": 20,  # Fecha Creación
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Llenar con datos
        for transporte in transportes:
            # Convertir códigos de choices a valores legibles
            estado = ESTADOS_DISPLAY.get(transporte["estado"], transporte["estado"])
            mes = MESES_DISPLAY.get(transporte["mes"], transporte["mes"])

            ws.append(
                [
                    estado,
                    mes,
                    transporte["transporte"],
                    transporte["cantidad"],
                    (
                        transporte["created_at"].strftime("%Y-%m-%d")
                        if transporte["created_at"]
                        else ""
                    ),
                ]
            )

        # Aplicar formato a las celdas numéricas
        for row in ws.iter_rows(min_row=4, max_col=4, max_row=ws.max_row):
            row[3].number_format = "#,##0"  # Formato para cantidad

        # Preparar la respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Reporte_Transportes.xlsx"

        return FileResponse(
            output,
            as_attachment=True,
            filename=filename,
            headers={
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            },
        )
