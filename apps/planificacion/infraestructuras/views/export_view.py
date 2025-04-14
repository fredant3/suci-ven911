from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from planificacion.infraestructuras.models import Infraestructura


class InfraestructuraExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "planificacion.listar_infraestructura"

    def get(self, request, *args, **kwargs):
        # Obtener todos los registros de infraestructura
        infraestructuras = Infraestructura.objects.all().values(
            "estado",
            "mes",
            "infraestructura",
            "cantidad",
            "created_at",
        )

        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Infraestructuras"

        # Configurar el título del reporte
        ws.merge_cells("A1:H1")
        title_cell = ws["A1"]
        title_cell.value = "Reporte de Infraestructuras Programadas"
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )

        # Espacio en blanco
        ws.append([])

        # Encabezados de columnas
        headers = [
            "Estado",
            "Mes",
            "Infraestructura",
            "Cantidad",
            "Fecha Creación",
        ]

        ws.append(headers)

        # Estilo para encabezados
        header_fill = PatternFill(
            start_color="DCE6F1", end_color="DCE6F1", fill_type="solid"
        )
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = Font(bold=True, color="4F81BD")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Configurar anchos de columnas
        column_widths = {
            "A": 15,  # Estado
            "B": 15,  # Mes
            "C": 35,  # Infraestructura
            "D": 15,  # Cantidad
            "E": 18,  # Fecha Creación
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Formato para números
        number_format = "#,##0"

        # Llenar con datos
        for infra in infraestructuras:
            row = [
                infra["estado"],
                infra["mes"],
                infra["infraestructura"],
                infra["cantidad"],
                (
                    infra["created_at"].strftime("%Y-%m-%d")
                    if infra["created_at"]
                    else ""
                ),
            ]
            ws.append(row)

            # Formatear columna de cantidad
            ws.cell(row=ws.max_row, column=4).number_format = number_format

        # Congelar los encabezados
        ws.freeze_panes = "A3"

        # Auto-filtro
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{ws.max_row}"

        # Preparar la respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="Reporte_Infraestructuras.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
