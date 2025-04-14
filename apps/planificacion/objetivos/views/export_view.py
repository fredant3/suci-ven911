from io import BytesIO
from planificacion.objetivos.models import Objetivo
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


class ObjetivoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "planificacion.listar_objetivo"

    def get(self, request, *args, **kwargs):
        # Obtener todos los objetivos con los campos necesarios
        objetivos = Objetivo.objects.all().values(
            "fechai",
            "fechaf",
            "objetiv",
            "meta",
            "created_at",
        )

        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active

        # Configurar el título del reporte
        ws.merge_cells("A1:E1")
        ws["A1"] = "Reporte de Objetivos"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF")
        ws.append([])  # Espacio en blanco

        # Encabezados de columnas
        ws.append(
            [
                "Fecha Inicio",
                "Fecha Fin",
                "Objetivo",
                "Meta",
                "Fecha Creación",
            ]
        )

        # Configurar anchos de columnas
        column_widths = {
            "A": 15,  # Fecha Inicio
            "B": 15,  # Fecha Fin
            "C": 40,  # Objetivo
            "D": 40,  # Meta
            "E": 15,  # Fecha Creación
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Llenar con datos
        for objetivo in objetivos:
            ws.append(
                [
                    (
                        objetivo["fechai"].strftime("%Y-%m-%d")
                        if objetivo["fechai"]
                        else ""
                    ),
                    (
                        objetivo["fechaf"].strftime("%Y-%m-%d")
                        if objetivo["fechaf"]
                        else ""
                    ),
                    objetivo["objetiv"],
                    objetivo["meta"],
                    (
                        objetivo["created_at"].strftime("%Y-%m-%d")
                        if objetivo["created_at"]
                        else ""
                    ),
                ]
            )

        # Preparar la respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="Objetivos.xlsx",
        )
