from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from planificacion.actividades.models import Actividad


class ActividadExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "planificacion.listar_actividad"

    def get(self, request, *args, **kwargs):
        # Obtener todas las actividades con los campos necesarios
        actividades = Actividad.objects.all().values(
            "fechai",
            "fechaf",
            "objetiv",
            "meta",
            "created_at",
        )

        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active

        # Configurar el título del reporte
        ws.merge_cells("A1:H1")
        ws["A1"] = "Reporte de Actividades"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF", size=14)
        ws.append([])  # Espacio en blanco

        # Encabezados de columnas
        headers = [
            "Fecha Inicio",
            "Fecha Fin",
            "Objetivo",
            "Meta",
            "Fecha Creación",
        ]

        ws.append(headers)

        # Estilo para encabezados
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Configurar anchos de columnas
        column_widths = {
            "A": 15,  # Fecha Inicio
            "B": 15,  # Fecha Fin
            "C": 40,  # Objetivo
            "D": 40,  # Meta
            "E": 18,  # Fecha Creación
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Llenar con datos
        for actividad in actividades:
            ws.append(
                [
                    (
                        actividad["fechai"].strftime("%Y-%m-%d")
                        if actividad["fechai"]
                        else ""
                    ),
                    (
                        actividad["fechaf"].strftime("%Y-%m-%d")
                        if actividad["fechaf"]
                        else ""
                    ),
                    actividad["objetiv"],
                    actividad["meta"],
                    (
                        actividad["created_at"].strftime("%Y-%m-%d")
                        if actividad["created_at"]
                        else ""
                    ),
                ]
            )

        # Congelar los encabezados
        ws.freeze_panes = "A3"

        # Preparar la respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="Reporte_Actividades.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
