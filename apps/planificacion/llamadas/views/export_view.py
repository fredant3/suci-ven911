from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from planificacion.llamadas.models import Llamada


class LlamadaExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "llamadas.listar_llamada"

    def get(self, request, *args, **kwargs):
        # Obtener todas las llamadas con los campos necesarios
        llamadas = Llamada.objects.all().values(
            "mes",
            "estado",
            "informativa",
            "falsa",
            "realesno",
            "realesf",
            "videop",
            "created_at",
        )

        # Crear el libro de Excel
        wb = Workbook()
        ws = wb.active

        # Configurar el título del reporte
        ws.merge_cells("A1:H1")
        ws["A1"] = "Reporte de Llamadas"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF")
        ws.append([])  # Espacio en blanco

        # Encabezados de columnas
        ws.append(
            [
                "Mes",
                "Estado",
                "Llamadas Informativas",
                "Llamadas Falsas",
                "Llamadas Reales No Atendidas",
                "Llamadas Reales Finalizadas",
                "Videollamadas Protección",
                "Fecha Creación",
            ]
        )

        # Configurar anchos de columnas
        column_widths = {
            "A": 10,  # Mes
            "B": 15,  # Estado
            "C": 20,  # Informativas
            "D": 15,  # Falsas
            "E": 25,  # Reales No Atendidas
            "F": 25,  # Reales Finalizadas
            "G": 20,  # Videollamadas
            "H": 15,  # Fecha Creación
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Llenar con datos
        for llamada in llamadas:
            ws.append(
                [
                    llamada["mes"],
                    llamada["estado"],
                    llamada["informativa"],
                    llamada["falsa"],
                    llamada["realesno"],
                    llamada["realesf"],
                    llamada["videop"],
                    (
                        llamada["created_at"].strftime("%Y-%m-%d")
                        if llamada["created_at"]
                        else ""
                    ),
                ]
            )

        # Preparar la respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="Reporte_Llamadas.xlsx",
        )
