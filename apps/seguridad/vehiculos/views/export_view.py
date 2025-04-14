from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from seguridad.vehiculos.models import Vehiculo


class VehiculoExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "vehiculo.listar_vehiculos"

    def get(self, request, *args, **kwargs):
        # Consulta optimizada con todos los campos relevantes
        vehiculos = Vehiculo.objects.filter(deleted_at__isnull=True).values(
            "nombre",
            "apellido",
            "cedula",
            "modelo",
            "vehiculo",
            "motivo",
            "capagasolina",
            "cantigasolina",
            "placa",
            "fecha",
            "hora",
            "created_at"
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Registro de Vehículos"

        # Configuración del título (ajustado a 12 columnas)
        ws.merge_cells("A1:L1")
        title_cell = ws["A1"]
        title_cell.value = "REGISTRO DE VEHÍCULOS"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados con anchos personalizados
        headers = [
            ("Nombre", 20),
            ("Apellido", 20),
            ("Cédula", 15),
            ("Modelo", 20),
            ("Tipo Vehículo", 20),
            ("Motivo", 25),
            ("Cap. Gasolina", 15),
            ("Cant. Gasolina", 15),
            ("Placa", 15),
            ("Fecha", 15),
            ("Hora", 15),
            ("Fecha Registro", 20)
        ]

        # Configurar columnas
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num) if col_num <= 26 else 'A' + chr(64 + col_num - 26)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Llenar datos
        for row_num, vehiculo in enumerate(vehiculos, 3):
            ws.append([
                vehiculo["nombre"],
                vehiculo["apellido"],
                vehiculo["cedula"],
                vehiculo["modelo"],
                vehiculo["vehiculo"],
                vehiculo["motivo"],
                vehiculo["capagasolina"],
                vehiculo["cantigasolina"],
                vehiculo["placa"],
                vehiculo["fecha"].strftime("%d/%m/%Y") if vehiculo["fecha"] else "N/A",
                vehiculo["hora"],
                vehiculo["created_at"].strftime("%d/%m/%Y %H:%M") if vehiculo["created_at"] else "N/A"
            ])

        # Aplicar bordes a todas las celdas
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="reporte_vehiculos.xlsx"
        )