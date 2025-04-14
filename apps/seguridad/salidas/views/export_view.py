from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from seguridad.salidas.models import Salida


class SalidaExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "salida.listar_salidas"

    def get(self, request, *args, **kwargs):
        # Consulta optimizada
        salidas = Salida.objects.filter(deleted_at__isnull=True).values(
            "name",
            "apellido",
            "cedula",
            "telefono",
            "fecha",
            "direccion",
            "cargo",
            "hora",
            "created_at"
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Registro de Salidas"

        # Configuración del título
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = "REGISTRO DE SALIDAS"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados (ajustados a 9 columnas)
        headers = [
            ("Nombre", 20),
            ("Apellido", 20),
            ("Cédula", 15),
            ("Teléfono", 15),
            ("Fecha", 15),
            ("Dirección", 30),
            ("Cargo", 20),
            ("Hora de Salida", 15),  # Cambiado de "Entrada" a "Salida"
            ("Fecha Registro", 20)
        ]

        # Configurar columnas
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Llenar datos
        for row_num, salida in enumerate(salidas, 3):
            ws.append([
                salida["name"],
                salida["apellido"],
                salida["cedula"],
                salida["telefono"] or "N/A",
                salida["fecha"].strftime("%d/%m/%Y") if salida["fecha"] else "N/A",
                salida["direccion"],
                salida["cargo"],
                salida["hora"],
                salida["created_at"].strftime("%d/%m/%Y %H:%M") if salida["created_at"] else "N/A"
            ])

        # Aplicar bordes
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="reporte_salidas.xlsx"
        )