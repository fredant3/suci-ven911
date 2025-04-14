from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from seguridad.entradas.models import Entrada


class EntradaExcelView(LoginRequiredMixin, CheckPermisosMixin, TemplateView):
    permission_required = "entrada.listar_entradas"

    def get(self, request, *args, **kwargs):
        # Consulta todos los registros no eliminados
        entradas = Entrada.objects.filter(deleted_at__isnull=True).values(
            "name",
            "apellido",
            "cedula",
            "telefono",
            "fecha",
            "direccion",
            "cargo",
            "hora",
            "created_at"
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Registro de Entradas"

        # Configuración del título
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = "REGISTRO DE ENTRADAS"
        title_cell.alignment = Alignment(horizontal="center")
        title_cell.font = Font(bold=True, size=14, color="0047AB")

        # Encabezados
        headers = [
            ("Nombre", 20),
            ("Apellido", 20),
            ("Cédula", 15),
            ("Teléfono", 15),
            ("Fecha", 15),
            ("Dirección", 30),
            ("Cargo", 20),
            ("Hora de Entrada", 15),
            ("Fecha Registro", 20)
        ]

        # Configurar columnas
        for col_num, (header, width) in enumerate(headers, 1):
            col_letter = chr(64 + col_num)
            ws.column_dimensions[col_letter].width = width
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Llenar datos
        for row_num, entrada in enumerate(entradas, 3):
            ws.append([
                entrada["name"],
                entrada["apellido"],
                entrada["cedula"],
                entrada["telefono"] or "N/A",
                entrada["fecha"].strftime("%d/%m/%Y") if entrada["fecha"] else "N/A",
                entrada["direccion"],
                entrada["cargo"],
                entrada["hora"],
                entrada["created_at"].strftime("%d/%m/%Y %H:%M") if entrada["created_at"] else "N/A"
            ])

        # Aplicar bordes
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.border = border

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="reporte_entradas.xlsx"
        )