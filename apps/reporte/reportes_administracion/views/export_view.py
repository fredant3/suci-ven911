from io import BytesIO
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import FileResponse
from django.views.generic import TemplateView
from helpers.CheckPermisosMixin import CheckPermisosMixin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# from administracion.averia.models import Averia  # Importar el modelo Averia
from apps.reporte.reportes_administracion.models import ReportesAdministracion


class ReportesAdministracionExcelView(
    LoginRequiredMixin, CheckPermisosMixin, TemplateView
):
    permission_required = (
        "reporte.listar_reportesadministracion"  # Usar el permiso correspondiente
    )

    def get(self, request, *args, **kwargs):
        reportesadministracion = ReportesAdministracion.objects.all().values(
            "tipo_averia__nombre",  # Campos del modelo con relaciones
            "departamento__nombre",
            "problema",
            "ubicacion",
            "serial",
            "codigo_bn",
            "created_at",
        )

        wb = Workbook()
        ws = wb.active

        # Configuración del título
        ws.merge_cells("A1:G1")
        ws["A1"] = "Reporte de Averías"
        ws["A1"].alignment = Alignment(horizontal="center")
        ws["A1"].font = Font(bold=True, color="0000FF")
        ws.append([])  # Espacio en blanco

        # Encabezados
        ws.append(
            [
                "Tipo de Avería",
                "Departamento",
                "Problema",
                "Ubicación",
                "Serial",
                "Código BN",
                "Fecha de Creación",
            ]
        )

        # Ajustes de columnas
        column_widths = {
            "A": 25,  # Tipo de Avería
            "B": 25,  # Departamento
            "C": 40,  # Problema
            "D": 30,  # Ubicación
            "E": 20,  # Serial
            "F": 20,  # Código BN
            "G": 20,  # Fecha
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Datos
        for reportesadministracion in ReportesAdministracion:
            fecha = (
                reportesadministracion["created_at"].strftime("%Y-%m-%d")
                if reportesadministracion["created_at"]
                else ""
            )
            ws.append(
                [
                    reportesadministracion["tipo_averia__nombre"],
                    reportesadministracion["departamento__nombre"],
                    reportesadministracion["problema"],
                    reportesadministracion["ubicacion"],
                    reportesadministracion["serial"],
                    reportesadministracion["codigo_bn"],
                    fecha,
                ]
            )

        # Generar respuesta
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return FileResponse(
            output,
            as_attachment=True,
            filename="reporte_administracion.xlsx",  # Nombre del archivo en minúsculas
        )
